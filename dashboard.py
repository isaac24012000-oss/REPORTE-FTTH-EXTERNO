import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
import os

st.set_page_config(
    page_title="Reporte Bitel FTTH",
    page_icon="📞",
    layout="wide",
    initial_sidebar_state="expanded"
)
# Actualizado 06/05/2026 - Preparado para MAYO 2026

# ============= CARGA DE DATOS DEL EXCEL =============

@st.cache_data(ttl=3600)  # 1 hora en local, más en Cloud
def load_mantra_data():
    """Carga datos de la hoja MANTRA del archivo REPORTE FTTH.xlsx
    Actualizado: 02/03/2026 - Ahora filtra por MES en lugar de FECHA"""
    excel_path = os.path.join(os.path.dirname(__file__), 'REPORTE FTTH.xlsx')
    
    try:
        df_mantra = pd.read_excel(excel_path, sheet_name='MANTRA')
        return df_mantra
    except Exception as e:
        return None

@st.cache_data(ttl=3600)    
def get_total_leads_and_conversion(mes_seleccionado="Noviembre"):
    """Obtiene total de leads y conversión para un mes específico"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 6589, 299  # Valores por defecto si no hay datos
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0, 0
    
    # Limpiar espacios en blanco
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    df_mes['NIVEL 3'] = df_mes['NIVEL 3'].astype(str).str.strip()
    
    # Total de leads para ese mes
    total_leads = len(df_mes)
    
    # Conversión: Con Cobertura + Contrato OK para ese mes
    df_conversion = df_mes[
        (df_mes['NIVEL 2'] == 'Con Cobertura') & 
        (df_mes['NIVEL 3'] == 'Contrato OK')
    ]
    total_conversion = len(df_conversion)
    
    return total_leads, total_conversion

@st.cache_data(ttl=60)  # Cache de 60 segundos para desarrollo
def get_conversion_mantra_mes(mes_seleccionado="Noviembre"):
    """Calcula la conversión: Ventas Del Mes (DRIVE) / Con Cobertura (MANTRA)
    = Total de Transacciones en DRIVE / Registros con cobertura en MANTRA"""
    df_mantra = load_mantra_data()
    df_drive = load_drive_data()
    
    if df_mantra is None or df_mantra.empty or df_drive is None or df_drive.empty:
        return 0
    
    # Obtener Con Cobertura de MANTRA para el mes
    df_mes_mantra = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    df_mes_mantra['NIVEL 2'] = df_mes_mantra['NIVEL 2'].astype(str).str.strip()
    con_cobertura = len(df_mes_mantra[df_mes_mantra['NIVEL 2'] == 'Con Cobertura'])
    
    if con_cobertura == 0:
        return 0
    
    # Obtener Ventas Del Mes del DRIVE (TODAS las transacciones, no solo INSTALADO)
    ventas_del_mes = get_ventas_del_mes_por_fecha(mes_seleccionado)
    
    if ventas_del_mes == 0:
        return 0
    
    # Conversión = Ventas Del Mes / Con Cobertura
    conversion_pct = round((ventas_del_mes / con_cobertura * 100)) if con_cobertura > 0 else 0
    return conversion_pct

@st.cache_data(ttl=3600)
def get_con_cobertura_count(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'Con Cobertura' para un mes específico"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Contar "Con Cobertura"
    con_cobertura = len(df_mes[df_mes['NIVEL 2'] == 'Con Cobertura'])
    
    return con_cobertura

@st.cache_data(ttl=3600)
def get_cancelados_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de cancelados para un mes específico usando columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    if mes_num is None:
        return 0
    
    # Filtrar por MES column si existe, sino por FECHA
    if 'MES' in df_drive.columns:
        df_mes = df_drive[
            (df_drive['MES'] == mes_seleccionado) &
            (df_drive['ESTADO'] == 'CANCELADO')
        ]
    else:
        # Fallback a FECHA
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        # Para Noviembre, incluir Octubre + Noviembre
        if mes_num == 11:
            df_mes = df_drive[
                ((df_drive['FECHA'].dt.month == 10) | (df_drive['FECHA'].dt.month == 11)) &
                (df_drive['ESTADO'] == 'CANCELADO')
            ]
        else:
            df_mes = df_drive[
                (df_drive['FECHA'].dt.month == mes_num) &
                (df_drive['ESTADO'] == 'CANCELADO')
            ]
    
    cancelados = len(df_mes)
    return cancelados

@st.cache_data
def get_instaladas_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de instaladas para un mes específico
    Regla: Solo INSTALADO (no incluye PENDIENTE)
    Filtra por columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    if mes_num is None:
        return 0
    
    # Para Noviembre, incluir Octubre + Noviembre
    es_noviembre = mes_num == 11
    instaladas = count_instaladas_con_regla(df_drive, mes_num, es_noviembre, mes_seleccionado)
    
    return instaladas

def get_ventas_generales_mes(mes_seleccionado="Noviembre"):
    """Obtiene el total de TODAS las transacciones del mes
    = INSTALADAS + PENDIENTES + CANCELADAS
    Filtra por columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Filtrar por mes usando columna MES
    if 'MES' in df_drive.columns:
        df_mes = df_drive[df_drive['MES'] == mes_seleccionado]
    else:
        # Fallback a FECHA si MES no existe
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_seleccionado, None)
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        df_mes = df_drive[df_drive['FECHA'].dt.month == mes_num]
    
    # Total de TODAS las transacciones
    total_transacciones = len(df_mes)
    return total_transacciones

@st.cache_data(ttl=60)  # Cache de 60 segundos para desarrollo
def get_ventas_del_mes_por_fecha(mes_seleccionado="Abril"):
    """Obtiene las transacciones para el mes especificado usando la columna MES del DRIVE.
    Esto es más confiable que filtrar por fechas ya que MES contiene el mes exacto."""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Mapeo de meses a nombres de columna MES
    mes_nombres = {
        'Enero': 'Enero', 'Febrero': 'Febrero', 'Marzo': 'Marzo', 'Abril': 'Abril',
        'Mayo': 'Mayo', 'Junio': 'Junio', 'Julio': 'Julio', 'Agosto': 'Agosto',
        'Septiembre': 'Septiembre', 'Octubre': 'Octubre', 'Noviembre': 'Noviembre', 'Diciembre': 'Diciembre'
    }
    
    mes_limpio = mes_nombres.get(mes_seleccionado, None)
    if mes_limpio is None:
        return 0
    
    # Filtrar por la columna MES directamente
    df_mes = df_drive[df_drive['MES'] == mes_limpio]
    
    # Contar todas las transacciones de ese mes
    total_ventas = len(df_mes)
    return total_ventas

@st.cache_data(ttl=3600)
def get_no_pago_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de NO PAGO para un mes específico usando columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    if mes_num is None:
        return 0
    
    # Limpiar espacios en blanco en MOTIVO CANCELACIÓN
    df_drive['MOTIVO CANCELACIÓN'] = df_drive['MOTIVO CANCELACIÓN'].astype(str).str.strip()
    
    # Filtrar por MES column si existe, sino por FECHA
    if 'MES' in df_drive.columns:
        df_mes = df_drive[
            (df_drive['MES'] == mes_seleccionado) &
            (df_drive['MOTIVO CANCELACIÓN'] == 'NO PAGO')
        ]
    else:
        # Fallback a FECHA
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        # Para Noviembre, incluir Octubre + Noviembre
        if mes_num == 11:
            df_mes = df_drive[
                ((df_drive['FECHA'].dt.month == 10) | (df_drive['FECHA'].dt.month == 11)) &
                (df_drive['MOTIVO CANCELACIÓN'] == 'NO PAGO')
            ]
        else:
            df_mes = df_drive[
                (df_drive['FECHA'].dt.month == mes_num) &
                (df_drive['MOTIVO CANCELACIÓN'] == 'NO PAGO')
            ]
    
    no_pago = len(df_mes)
    return no_pago

@st.cache_data(ttl=3600)
def get_no_responde_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'No Responde' para un mes específico desde MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco en NIVEL 1
    df_mes['NIVEL 1'] = df_mes['NIVEL 1'].astype(str).str.strip()
    
    # Contar "No Responde"
    no_responde = len(df_mes[df_mes['NIVEL 1'] == 'No Responde'])
    
    return no_responde

@st.cache_data(ttl=3600)
def get_no_especifica_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'No Especifica' para un mes específico desde MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco en NIVEL 2
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Contar "No Especifica"
    no_especifica = len(df_mes[df_mes['NIVEL 2'] == 'No Especifica'])
    
    return no_especifica

@st.cache_data(ttl=3600)
def get_sin_cobertura_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'Sin Cobertura' para un mes específico desde MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco en NIVEL 2
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Contar "Sin Cobertura"
    sin_cobertura = len(df_mes[df_mes['NIVEL 2'] == 'Sin Cobertura'])
    
    return sin_cobertura

@st.cache_data(ttl=3600)
def load_lista_metas():
    """Carga los datos de metas por mes de la hoja LISTA"""
    excel_path = os.path.join(os.path.dirname(__file__), 'REPORTE FTTH.xlsx')
    
    try:
        df_lista = pd.read_excel(excel_path, sheet_name='LISTA')
        return df_lista
    except Exception as e:
        return None

@st.cache_data(ttl=60)  # Reducido a 60 segundos para detectar cambios rápidamente
def get_meses_disponibles():
    """Obtiene dinámicamente todos los meses disponibles en DRIVE (hoja de ventas)
    Retorna lista de tuplas (mes_año, mes_nombre, año, mes_num) ordenadas por mes y año"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        # Fallback a los meses por defecto
        return [
            ('Noviembre 2025', 'Noviembre', 2025, 11),
            ('Diciembre 2025', 'Diciembre', 2025, 12),
            ('Enero 2026', 'Enero', 2026, 1),
            ('Febrero 2026', 'Febrero', 2026, 2),
            ('Marzo 2026', 'Marzo', 2026, 3),
            ('Abril 2026', 'Abril', 2026, 4)
        ]
    
    # Obtener meses únicos del DRIVE
    meses_unicos = df_drive['MES'].dropna().unique().tolist()
    
    if not meses_unicos:
        return [
            ('Noviembre 2025', 'Noviembre', 2025, 11),
            ('Diciembre 2025', 'Diciembre', 2025, 12),
            ('Enero 2026', 'Enero', 2026, 1),
            ('Febrero 2026', 'Febrero', 2026, 2),
            ('Marzo 2026', 'Marzo', 2026, 3),
            ('Abril 2026', 'Abril', 2026, 4)
        ]
    
    # Mapeo de meses a números
    mes_num_map = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    
    # Orden correcto de meses
    orden_meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    # Filtrar y ordenar meses según el orden correcto
    meses_ordenados = [m for m in orden_meses if m in meses_unicos]
    
    # Si hay meses no en el orden (edge case), agregarlos al final
    meses_faltantes = [m for m in meses_unicos if m not in orden_meses]
    meses_ordenados.extend(meses_faltantes)
    
    # Crear tuplas (mes_año, mes_nombre, año, mes_num)
    # Asumir año basado en si es antes o después de junio
    resultado = []
    for mes_nombre in meses_ordenados:
        mes_num = mes_num_map.get(mes_nombre, 1)
        # Si es Noviembre-Diciembre, asumir 2025; si es Enero-Octubre, asumir 2026
        año = 2025 if mes_num >= 11 else 2026
        mes_año = f"{mes_nombre} {año}"
        resultado.append((mes_año, mes_nombre, año, mes_num))
    
    return resultado

@st.cache_data(ttl=60)
def get_mes_mas_reciente():
    """Obtiene el mes más reciente disponible en DRIVE basado en fechas
    Retorna tupla (mes_año, mes_nombre, año, mes_num)"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return ('Abril 2026', 'Abril', 2026, 4)
    
    # Convertir FECHA a datetime
    df_temp = df_drive.copy()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    df_temp = df_temp[df_temp['FECHA'].notna()]
    
    if df_temp.empty:
        return ('Abril 2026', 'Abril', 2026, 4)
    
    # Obtener el mes y año más reciente
    fecha_max = df_temp['FECHA'].max()
    mes_num_reciente = fecha_max.month
    año_reciente = fecha_max.year
    
    # Mapeo de número de mes a nombre
    mes_num_a_nombre = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    mes_nombre_reciente = mes_num_a_nombre.get(mes_num_reciente, 'Abril')
    mes_año_reciente = f"{mes_nombre_reciente} {año_reciente}"
    
    return (mes_año_reciente, mes_nombre_reciente, año_reciente, mes_num_reciente)

@st.cache_data(ttl=3600)  # 1 hora de caché
def load_drive_data():
    """Carga datos de la hoja DRIVE del archivo REPORTE FTTH.xlsx"""
    excel_path = os.path.join(os.path.dirname(__file__), 'REPORTE FTTH.xlsx')
    
    try:
        df_drive = pd.read_excel(excel_path, sheet_name='DRIVE')
        return df_drive
    except Exception as e:
        return None

def count_instaladas_con_regla(df, fecha_mes_num, fecha_mes_es_noviembre=False, mes_nombre="Enero"):
    """
    Cuenta instaladas aplicando regla para todos los meses.
    
    Regla de VENTAS INSTALADAS DEL MES:
    - Solo INSTALADO
    - Sin considerar PENDIENTE
    - Filtra por columna MES (no por FECHA)
    
    Fórmula: COUNT(ESTADO='INSTALADO')
    
    Args:
        df: DataFrame del DRIVE
        fecha_mes_num: número del mes (deprecated, usa mes_nombre)
        fecha_mes_es_noviembre: si incluir Oct+Nov (deprecated)
        mes_nombre: nombre del mes ('Enero', 'Diciembre', etc)
    
    Returns:
        int: cantidad de instaladas según la regla
    """
    # Preparar dataframe
    df_temp = df.copy()
    df_temp['ESTADO'] = df_temp['ESTADO'].astype(str).str.strip()
    
    # Filtrar por columna MES (no por FECHA)
    if 'MES' in df_temp.columns:
        df_mes = df_temp[df_temp['MES'] == mes_nombre]
    else:
        # Fallback a filtro por FECHA si MES no existe
        df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
        if fecha_mes_es_noviembre:
            df_mes = df_temp[
                ((df_temp['FECHA'].dt.month == 10) | (df_temp['FECHA'].dt.month == 11))
            ]
        else:
            df_mes = df_temp[df_temp['FECHA'].dt.month == fecha_mes_num]
    
    # Aplicar regla: Solo INSTALADO
    df_instaladas = df_mes[df_mes['ESTADO'] == 'INSTALADO']
    
    return len(df_instaladas)

@st.cache_data(ttl=3600)
def debug_instaladas_por_dia(mes_seleccionado="Febrero", dia=3):
    """Función de debug para ver qué registros hay en un día específico"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return None
    
    df_temp = df_drive.copy()
    
    # Sin limpiezas, solo conversion
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # Filtrar por mes y día
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    df_temp['FECHA_MES'] = df_temp['FECHA'].dt.month
    df_temp['FECHA_DIA'] = df_temp['FECHA'].dt.day
    
    # Filtrar por mes y día
    df_filtrado = df_temp[(df_temp['FECHA_MES'] == mes_num) & (df_temp['FECHA_DIA'] == dia)]
    
    # Retornar TODOS los registros sin filtrar
    return df_filtrado

@st.cache_data(ttl=60)  # Reducido a 60 segundos para datos actualizados de Abril
def get_instaladas_por_semana(mes_seleccionado="Noviembre"):
    """Obtiene VENTAS por DÍA para un mes específico.
    VENTAS = todos los registros del mes (sin importar PAGO o ESTADO)
    Retorna un DataFrame con día y cantidad de ventas
    Filtra por fecha actual para no mostrar registros futuros"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    # Conversión simple, sin limpieza excesiva
    df_temp = df_drive.copy()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # Mapeo de meses
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    
    mes_num = mes_numeros.get(mes_seleccionado, None)
    if mes_num is None:
        return pd.DataFrame()
    
    # Filtrar por fecha válida
    df_temp = df_temp[df_temp['FECHA'].notna()]
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Extraer mes y año de FECHA
    df_temp['FECHA_MES'] = df_temp['FECHA'].dt.month
    df_temp['FECHA_AÑO'] = df_temp['FECHA'].dt.year
    df_temp['FECHA_DIA'] = df_temp['FECHA'].dt.day
    
    # Filtrar por mes exacto
    df_mes = df_temp[df_temp['FECHA_MES'] == mes_num].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Si hay múltiples años, tomar el más reciente
    año_filtro = df_mes['FECHA_AÑO'].max()
    df_mes = df_mes[df_mes['FECHA_AÑO'] == año_filtro]
    
    # Filtrar VENTAS - todos los registros sin importar PAGO o ESTADO
    df_ventas = df_mes.copy()
    
    if df_ventas.empty:
        return pd.DataFrame()
    
    # Validar días válidos del mes
    if mes_num == 12:
        último_día_mes = pd.Timestamp(year=año_filtro+1, month=1, day=1) - pd.DateOffset(days=1)
    else:
        último_día_mes = pd.Timestamp(year=año_filtro, month=mes_num+1, day=1) - pd.DateOffset(days=1)
    
    último_día_válido = último_día_mes.day
    
    # Filtrar días válidos
    df_ventas = df_ventas[(df_ventas['FECHA_DIA'] >= 1) & (df_ventas['FECHA_DIA'] <= último_día_válido)]
    
    if df_ventas.empty:
        return pd.DataFrame()
    
    # Contar por día
    df_dias = df_ventas.groupby('FECHA_DIA').size().reset_index(name='INSTALADAS')
    df_dias.columns = ['DIA', 'INSTALADAS']
    
    # Crear etiquetas
    mes_nombres_cortos = {
        1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
    }
    mes_str = mes_nombres_cortos[mes_num]
    
    df_dias['DIA_ETIQUETA'] = df_dias['DIA'].astype(str) + ' ' + mes_str
    
    # Retornar ordenado por DIA (número), no por DIA_ETIQUETA (string)
    result = df_dias.sort_values('DIA')[['DIA_ETIQUETA', 'INSTALADAS']]
    result.columns = ['DIA', 'INSTALADAS']
    
    return result

@st.cache_data(ttl=3600)
def get_comparativo_semanas_multiples_meses():
    """Obtiene un comparativo de instaladas por DÍA para todos los meses disponibles.
    Retorna un DataFrame con día y cantidad por cada mes
    Filtra por fecha actual para no mostrar registros futuros"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    df_temp = df_drive.copy()
    df_temp['ESTADO'] = df_temp['ESTADO'].astype(str).str.strip()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Filtrar solo instaladas
    df_instaladas = df_temp[df_temp['ESTADO'] == 'INSTALADO'].copy()
    
    if df_instaladas.empty:
        return pd.DataFrame()
    
    # Obtener columna MES si existe, sino calcularla
    if 'MES' not in df_instaladas.columns:
        mes_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        df_instaladas['MES'] = df_instaladas['FECHA'].dt.month.map(mes_nombres)
    
    # Extraer día del mes
    df_instaladas['DIA'] = df_instaladas['FECHA'].dt.day
    
    # Agrupar por mes y día
    df_pivot = df_instaladas.groupby(['MES', 'DIA']).size().reset_index(name='INSTALADAS')
    
    # Crear tabla pivote (meses como columnas, días como filas)
    df_comparativo = df_pivot.pivot(index='DIA', columns='MES', values='INSTALADAS').fillna(0).astype(int)
    
    # Ordenar por día
    df_comparativo = df_comparativo.sort_index()
    
    return df_comparativo

def get_comparativo_acumulativo_multiples_meses():
    """Obtiene un comparativo ACUMULATIVO de instaladas para todos los meses disponibles.
    Retorna un DataFrame con día y cantidad acumulada por cada mes
    Filtra por fecha actual para no mostrar registros futuros"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    df_temp = df_drive.copy()
    df_temp['ESTADO'] = df_temp['ESTADO'].astype(str).str.strip()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Filtrar solo instaladas
    df_instaladas = df_temp[df_temp['ESTADO'] == 'INSTALADO'].copy()
    
    if df_instaladas.empty:
        return pd.DataFrame()
    
    # Obtener columna MES si existe, sino calcularla
    if 'MES' not in df_instaladas.columns:
        mes_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        df_instaladas['MES'] = df_instaladas['FECHA'].dt.month.map(mes_nombres)
    
    # Extraer día del mes
    df_instaladas['DIA'] = df_instaladas['FECHA'].dt.day
    
    # Agrupar por mes y día
    df_pivot = df_instaladas.groupby(['MES', 'DIA']).size().reset_index(name='INSTALADAS')
    
    # Crear tabla pivote (meses como columnas, días como filas)
    df_comparativo = df_pivot.pivot(index='DIA', columns='MES', values='INSTALADAS').fillna(0).astype(int)
    
    # Ordenar por día
    df_comparativo = df_comparativo.sort_index()
    
    # Calcular acumulados para cada mes
    df_acumulativo = df_comparativo.cumsum()
    
    return df_acumulativo

@st.cache_data(ttl=3600)
def get_comparativo_diario_multiples_meses():
    """Obtiene un comparativo de VENTAS por DÍA para todos los meses disponibles (últimos 6 meses).
    VENTAS = todos los registros del mes (sin importar PAGO o ESTADO)
    Retorna un DataFrame con día como índice y cada mes como columna
    Filtra por fecha actual para no mostrar registros futuros"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    df_temp = df_drive.copy()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Extraer mes y año
    df_temp['FECHA_MES_NUM'] = df_temp['FECHA'].dt.month
    df_temp['FECHA_AÑO'] = df_temp['FECHA'].dt.year
    df_temp['FECHA_DIA'] = df_temp['FECHA'].dt.day
    
    mes_nombres = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    df_temp['MES_NOMBRE'] = df_temp['FECHA_MES_NUM'].map(mes_nombres)
    
    # Obtener meses únicos disponibles y ordenarlos cronológicamente
    df_temp['FECHA_SORT'] = df_temp['FECHA_AÑO'] * 100 + df_temp['FECHA_MES_NUM']
    meses_disponibles = df_temp.sort_values('FECHA_SORT')['MES_NOMBRE'].unique()
    
    # Tomar últimos 6 meses
    meses_a_mostrar = list(meses_disponibles[-6:])
    
    # Filtrar datos para estos meses
    df_filtrado = df_temp[df_temp['MES_NOMBRE'].isin(meses_a_mostrar)].copy()
    
    if df_filtrado.empty:
        return pd.DataFrame()
    
    # Agrupar por mes y día (contando todos los registros)
    df_pivot = df_filtrado.groupby(['MES_NOMBRE', 'FECHA_DIA']).size().reset_index(name='VENTAS')
    
    # Crear tabla pivote (días como filas, meses como columnas)
    df_comparativo = df_pivot.pivot(index='FECHA_DIA', columns='MES_NOMBRE', values='VENTAS').fillna(0).astype(int)
    
    # Reordenar columnas en orden cronológico
    df_comparativo = df_comparativo[meses_a_mostrar]
    
    # Ordenar por día
    df_comparativo = df_comparativo.sort_index()
    
    # Agregar columna de Día
    df_comparativo.insert(0, 'Día', df_comparativo.index)
    df_comparativo['Día'] = df_comparativo['Día'].astype(int)
    
    return df_comparativo

@st.cache_data(ttl=3600)
def get_progreso_semanal(mes_seleccionado="Abril"):
    """Obtiene el progreso semanal vs la meta esperada.
    Meta esperada por semana: 
    - Semana 1 (días 1-7): 25% 
    - Semana 2 (días 8-14): 50%
    - Semana 3 (días 15-21): 75%
    - Semana 4 (días 22-31): 100%
    Retorna DataFrame con semana, instaladas reales, meta esperada, y porcentaje de cumplimiento"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    df_temp = df_drive.copy()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # Mapeo de meses
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    
    mes_num = mes_numeros.get(mes_seleccionado, None)
    if mes_num is None:
        return pd.DataFrame()
    
    # Filtrar por fecha válida
    df_temp = df_temp[df_temp['FECHA'].notna()]
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Extraer mes y año de FECHA
    df_temp['FECHA_MES'] = df_temp['FECHA'].dt.month
    df_temp['FECHA_AÑO'] = df_temp['FECHA'].dt.year
    df_temp['FECHA_DIA'] = df_temp['FECHA'].dt.day
    
    # Filtrar por mes exacto
    df_mes = df_temp[df_temp['FECHA_MES'] == mes_num].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Si hay múltiples años, tomar el más reciente
    año_filtro = df_mes['FECHA_AÑO'].max()
    df_mes = df_mes[df_mes['FECHA_AÑO'] == año_filtro]
    
    # Contar por día (todos los registros sin importar PAGO o ESTADO)
    df_dias = df_mes.groupby('FECHA_DIA').size().reset_index(name='INSTALADAS')
    
    # Calcular acumulado
    df_dias['ACUMULADO'] = df_dias['INSTALADAS'].cumsum()
    
    # Obtener total mensual
    total_mes = df_dias['INSTALADAS'].sum()
    
    # Asignar a semanas
    semanas_data = []
    
    # Semana 1: días 1-7
    df_semana1 = df_dias[df_dias['FECHA_DIA'] <= 7]
    acum_sem1 = df_semana1['INSTALADAS'].sum() if not df_semana1.empty else 0
    semanas_data.append({
        'Semana': 1,
        'Dias': '1-7',
        'Reales': acum_sem1,
        'Meta_Esperada': round(total_mes * 0.25) if total_mes > 0 else 0,
        'Total_Mes': total_mes
    })
    
    # Semana 2: días 8-14
    df_semana2 = df_dias[(df_dias['FECHA_DIA'] >= 8) & (df_dias['FECHA_DIA'] <= 14)]
    acum_sem2 = acum_sem1 + (df_semana2['INSTALADAS'].sum() if not df_semana2.empty else 0)
    semanas_data.append({
        'Semana': 2,
        'Dias': '8-14',
        'Reales': acum_sem2,
        'Meta_Esperada': round(total_mes * 0.50) if total_mes > 0 else 0,
        'Total_Mes': total_mes
    })
    
    # Semana 3: días 15-21
    df_semana3 = df_dias[(df_dias['FECHA_DIA'] >= 15) & (df_dias['FECHA_DIA'] <= 21)]
    acum_sem3 = acum_sem2 + (df_semana3['INSTALADAS'].sum() if not df_semana3.empty else 0)
    semanas_data.append({
        'Semana': 3,
        'Dias': '15-21',
        'Reales': acum_sem3,
        'Meta_Esperada': round(total_mes * 0.75) if total_mes > 0 else 0,
        'Total_Mes': total_mes
    })
    
    # Semana 4: días 22-31
    df_semana4 = df_dias[df_dias['FECHA_DIA'] >= 22]
    acum_sem4 = acum_sem3 + (df_semana4['INSTALADAS'].sum() if not df_semana4.empty else 0)
    semanas_data.append({
        'Semana': 4,
        'Dias': '22-31',
        'Reales': acum_sem4,
        'Meta_Esperada': round(total_mes * 1.00) if total_mes > 0 else 0,
        'Total_Mes': total_mes
    })
    
    df_resultado = pd.DataFrame(semanas_data)
    
    # Calcular porcentaje de cumplimiento vs meta esperada
    df_resultado['Cumplimiento%'] = df_resultado.apply(
        lambda row: int((row['Reales'] / row['Meta_Esperada'] * 100)) if row['Meta_Esperada'] > 0 else 0,
        axis=1
    )
    
    return df_resultado

@st.cache_data(ttl=3600)
def get_semana_actual():
    """Determina qué semana es hoy dentro del mes actual"""
    hoy = pd.Timestamp.today()
    dia = hoy.day
    
    if dia <= 7:
        return 1, '1-7'
    elif dia <= 14:
        return 2, '8-14'
    elif dia <= 21:
        return 3, '15-21'
    else:
        return 4, '22-31'

@st.cache_data(ttl=3600)
def get_cumplimiento_asesor_semana_actual(asesor, meta_mensual, mes_seleccionado="Abril"):
    """Obtiene el cumplimiento del asesor contra su meta mensual.
    Calcula: (Ventas reales acumuladas en semana / Meta mensual) × 100
    
    Retorna el porcentaje de cumplimiento contra la meta mensual total.
    Evaluación basada en semana:
    - Semana 1: Debe estar al ≥25% de su meta
    - Semana 2: Debe estar al ≥50% de su meta
    - Semana 3: Debe estar al ≥75% de su meta
    - Semana 4: Debe estar al ≥100% de su meta
    """
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty or meta_mensual <= 0:
        return 0
    
    # Obtener semana actual
    semana_num, rango_dias = get_semana_actual()
    
    # Limpiar espacios en asesor
    df_drive['ASESOR'] = df_drive['ASESOR'].astype(str).str.strip()
    asesor_clean = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor_clean)
    
    # Convertir FECHA a datetime
    df_temp = df_drive.copy()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    df_temp = df_temp[df_temp['FECHA'].notna()]
    
    # Mapeo de meses
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    if mes_num is None:
        return 0
    
    # Filtra por mes y asesor (TODAS las transacciones sin importar ESTADO)
    df_temp['FECHA_MES'] = df_temp['FECHA'].dt.month
    df_temp['FECHA_DIA'] = df_temp['FECHA'].dt.day
    df_temp['ESTADO'] = df_temp['ESTADO'].astype(str).str.strip()
    
    df_asesor_mes = df_temp[
        (df_temp['FECHA_MES'] == mes_num) &
        (df_temp['ASESOR'].isin(nombres_alternativos))
    ]
    
    if df_asesor_mes.empty:
        return 0
    
    # Contar ventas acumuladas hasta hoy en la semana actual
    if semana_num == 1:
        ventas_semana = len(df_asesor_mes[df_asesor_mes['FECHA_DIA'] <= 7])
    elif semana_num == 2:
        ventas_semana = len(df_asesor_mes[df_asesor_mes['FECHA_DIA'] <= 14])
    elif semana_num == 3:
        ventas_semana = len(df_asesor_mes[df_asesor_mes['FECHA_DIA'] <= 21])
    else:  # semana 4
        ventas_semana = len(df_asesor_mes[df_asesor_mes['FECHA_DIA'] <= 31])
    
    # Calcular cumplimiento% contra la META MENSUAL (no semanal)
    cumplimiento_pct = int((ventas_semana / meta_mensual * 100)) if meta_mensual > 0 else 0
    
    return cumplimiento_pct

@st.cache_data(ttl=3600)
def get_nombres_alternativos(asesor):
    """Obtiene múltiples variantes del nombre del asesor para búsqueda flexible"""
    nombres = [asesor.strip()]
    # Agregar variante sin números al final (ej: ST2_VTP -> ST_VTP)
    import re
    nombre_sin_num = re.sub(r'(\d+)(_VTP)$', r'\2', asesor)
    if nombre_sin_num != asesor:
        nombres.append(nombre_sin_num)
    # Agregar variante con número (ej: ST_VTP -> ST2_VTP)
    nombre_con_num = re.sub(r'(_VTP)$', r'2_VTP', asesor.replace('2_VTP', '_VTP'))
    if nombre_con_num != asesor and '2_VTP' in nombre_con_num:
        nombres.append(nombre_con_num)
    return nombres

def get_pendientes_asesor_mes(asesor, mes_seleccionado="Enero"):
    """Obtiene cantidad de transacciones PENDIENTE por asesor para un mes"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_drive['ASESOR'] = df_drive['ASESOR'].astype(str).str.strip()
    asesor = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor)
    
    # Filtrar por mes y asesor (probando múltiples nombres)
    if 'MES' in df_drive.columns:
        df_mes_asesor = df_drive[(df_drive['MES'] == mes_seleccionado) & (df_drive['ASESOR'].isin(nombres_alternativos))]
    else:
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_seleccionado, None)
        if mes_num is None:
            return 0
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        df_mes_asesor = df_drive[(df_drive['FECHA'].dt.month == mes_num) & (df_drive['ASESOR'].isin(nombres_alternativos))]
    
    # Contar PENDIENTE
    df_mes_asesor['ESTADO'] = df_mes_asesor['ESTADO'].astype(str).str.strip()
    pendientes = len(df_mes_asesor[df_mes_asesor['ESTADO'] == 'PENDIENTE'])
    return pendientes

def get_leads_asesor_mes(asesor, mes_seleccionado="Enero"):
    """Obtiene el total de leads asignados a un asesor en un mes de MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_mantra['Agente'] = df_mantra['Agente'].astype(str).str.strip()
    asesor = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor)
    
    # Obtener datos del asesor en MANTRA para el mes
    df_mes_asesor = df_mantra[(df_mantra['Mes'] == mes_seleccionado) & (df_mantra['Agente'].isin(nombres_alternativos))]
    
    # Total de leads (registros)
    total_leads = len(df_mes_asesor)
    return total_leads

def get_con_cobertura_asesor_mes(asesor, mes_seleccionado="Enero"):
    """Obtiene la cantidad de leads con cobertura para un asesor en un mes"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_mantra['Agente'] = df_mantra['Agente'].astype(str).str.strip()
    asesor = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor)
    
    # Obtener datos del asesor en MANTRA para el mes
    df_mes_asesor = df_mantra[(df_mantra['Mes'] == mes_seleccionado) & (df_mantra['Agente'].isin(nombres_alternativos))]
    
    if df_mes_asesor.empty:
        return 0
    
    # Limpiar espacios en NIVEL 2
    df_mes_asesor['NIVEL 2'] = df_mes_asesor['NIVEL 2'].astype(str).str.strip()
    
    # Contar "Con Cobertura"
    con_cobertura = len(df_mes_asesor[df_mes_asesor['NIVEL 2'] == 'Con Cobertura'])
    return con_cobertura

@st.cache_data(ttl=3600)
def get_ventas_asesor_mes(asesor, mes_seleccionado="Enero"):
    """Obtiene el total de ventas (PAGO) del asesor en DRIVE para un mes específico"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_drive['ASESOR'] = df_drive['ASESOR'].astype(str).str.strip()
    asesor_clean = asesor.strip()
    
    # Filtrar por mes del DRIVE
    if 'MES' in df_drive.columns:
        df_mes_drive = df_drive[
            (df_drive['MES'] == mes_seleccionado) & 
            (df_drive['ASESOR'] == asesor_clean)
        ].copy()
    else:
        df_mes_drive = df_drive[df_drive['ASESOR'] == asesor_clean].copy()
    
    if df_mes_drive.empty:
        return 0
    
    # Contar transacciones con PAGO (cualquier valor no vacío en PAGO)
    df_mes_drive['PAGO'] = df_mes_drive['PAGO'].astype(str).str.strip()
    total_ventas = len(df_mes_drive[
        (df_mes_drive['PAGO'] != '') & 
        (df_mes_drive['PAGO'] != 'nan') &
        (df_mes_drive['PAGO'].notna())
    ])
    return total_ventas

def get_conversion_asesor_mes(asesor, mes_seleccionado="Noviembre"):
    """Calcula la conversión por asesor: Transacciones CON PAGO en DRIVE / Con Cobertura (de MANTRA)
    Cuenta todas las transacciones (INSTALADAS + CANCELADAS) que tengan PAGO"""
    df_drive = load_drive_data()
    df_mantra = load_mantra_data()
    
    if df_drive is None or df_drive.empty or df_mantra is None or df_mantra.empty:
        return 0
    
    # ========= CONTAR PAGO EN DRIVE =========
    # Filtrar DRIVE por mes y asesor
    df_drive_temp = df_drive.copy()
    df_drive_temp['ASESOR'] = df_drive_temp['ASESOR'].astype(str).str.strip()
    asesor_clean = asesor.strip()
    
    # Filtrar por mes del DRIVE
    if 'MES' in df_drive_temp.columns:
        df_mes_drive = df_drive_temp[
            (df_drive_temp['MES'] == mes_seleccionado) & 
            (df_drive_temp['ASESOR'] == asesor_clean)
        ].copy()
    else:
        df_mes_drive = df_drive_temp[df_drive_temp['ASESOR'] == asesor_clean].copy()
    
    if df_mes_drive.empty:
        return 0
    
    # Contar transacciones con PAGO (sin importar ESTADO)
    # PAGO = cualquier valor que NO sea null/NaN/vacío
    df_mes_drive['PAGO'] = df_mes_drive['PAGO'].astype(str).str.strip()
    transacciones_pago = len(df_mes_drive[
        (df_mes_drive['PAGO'] != '') & 
        (df_mes_drive['PAGO'] != 'nan') &
        (df_mes_drive['PAGO'].notna())
    ])
    
    # ========= CONTAR CON COBERTURA EN MANTRA =========
    # Limpiar espacios en los nombres de asesor en MANTRA
    df_mantra['Agente'] = df_mantra['Agente'].astype(str).str.strip()
    
    # Obtener nombres alternativos del asesor
    nombres_alternativos = get_nombres_alternativos(asesor_clean)
    
    # Obtener datos del asesor en MANTRA para el mes
    df_mes_mantra = None
    for nombre in nombres_alternativos:
        df_temp = df_mantra[(df_mantra['Mes'] == mes_seleccionado) & (df_mantra['Agente'] == nombre)].copy()
        if not df_temp.empty:
            df_mes_mantra = df_temp
            break
    
    if df_mes_mantra is None or df_mes_mantra.empty:
        return 0
    
    # Limpiar NIVEL 2
    df_mes_mantra['NIVEL 2'] = df_mes_mantra['NIVEL 2'].astype(str).str.strip()
    
    # Contar "Con Cobertura"
    con_cobertura = len(df_mes_mantra[df_mes_mantra['NIVEL 2'] == 'Con Cobertura'])
    
    if con_cobertura == 0:
        return 0
    
    # Conversión = Transacciones con PAGO / Con Cobertura
    conversion_pct = round((transacciones_pago / con_cobertura * 100)) if con_cobertura > 0 else 0
    return conversion_pct

@st.cache_data(ttl=3600)
def get_datos_mantra_mes(mes_seleccionado="Febrero"):
    """Obtiene datos detallados de MANTRA para un mes específico sin agregación"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Limpiar espacios en blanco
    df_mes['Agente'] = df_mes['Agente'].astype(str).str.strip()
    df_mes['NIVEL 1'] = df_mes['NIVEL 1'].astype(str).str.strip()
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    df_mes['NIVEL 3'] = df_mes['NIVEL 3'].astype(str).str.strip()
    
    return df_mes

@st.cache_data(ttl=3600)
def get_casos_por_agente_nivel(mes_seleccionado="Febrero"):
    """Obtiene casos por agente y nivel (1, 2, 3) desde MANTRA
    Retorna un DataFrame con información detallada para análisis"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Limpiar espacios en blanco
    df_mes['Agente'] = df_mes['Agente'].astype(str).str.strip()
    df_mes['NIVEL 1'] = df_mes['NIVEL 1'].astype(str).str.strip()
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    df_mes['NIVEL 3'] = df_mes['NIVEL 3'].astype(str).str.strip()
    
    # Agrupar por Agente
    agentes = df_mes['Agente'].unique()
    
    datos = []
    for agente in sorted(agentes):
        df_agente = df_mes[df_mes['Agente'] == agente]
        total_casos = len(df_agente)
        
        # Contar por NIVEL 1
        nivel1_counts = df_agente['NIVEL 1'].value_counts().to_dict()
        
        # Contar por NIVEL 2
        nivel2_counts = df_agente['NIVEL 2'].value_counts().to_dict()
        
        # Contar por NIVEL 3
        nivel3_counts = df_agente['NIVEL 3'].value_counts().to_dict()
        
        datos.append({
            'Agente': agente,
            'Total Casos': total_casos,
            'NIVEL 1': nivel1_counts,
            'NIVEL 2': nivel2_counts,
            'NIVEL 3': nivel3_counts
        })
    
    return pd.DataFrame(datos)

def calculate_drive_metrics(metas_dict, mes_filtro=None, mes_nombre=None):

    """
    Calcula Cumplimiento y Efectividad por asesor usando datos de DRIVE
    
    Cumplimiento = INSTALADAS / META
    Efectividad = INSTALADAS / (INSTALADAS + CANCELADAS)
    """
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return {}
    
    # Extraer columnas necesarias
    df_drive = df_drive[['FECHA', 'MES', 'ASESOR', 'ESTADO', 'PAGO']].copy()
    
    # Filtrar por mes usando columna MES si está disponible
    if mes_nombre and 'MES' in df_drive.columns:
        df_drive = df_drive[df_drive['MES'] == mes_nombre]
    elif mes_filtro:
        # Fallback a FECHA si MES no existe
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        df_drive = df_drive[df_drive['FECHA'].dt.month == mes_filtro]
    
    # Contar INSTALADOS por asesor (solo INSTALADO, sin PENDIENTE)
    df_drive_temp = df_drive.copy()
    df_drive_temp['ESTADO'] = df_drive_temp['ESTADO'].astype(str).str.strip()
    
    # Solo INSTALADO
    df_instalados = df_drive_temp[df_drive_temp['ESTADO'] == 'INSTALADO']
    instalados_por_asesor = df_instalados.groupby('ASESOR').size()
    
    # Contar CANCELADOS por asesor
    cancelados_por_asesor = df_drive[df_drive['ESTADO'] == 'CANCELADO'].groupby('ASESOR').size()
    
    # Calcular métricas
    metricas = {}
    for asesor, meta in metas_dict.items():
        instalados = instalados_por_asesor.get(asesor, 0)
        cancelados = cancelados_por_asesor.get(asesor, 0)
        
        # Cumplimiento = INSTALADAS / META
        cumplimiento = round((instalados / meta * 100) if meta > 0 else 0)
        
        # Efectividad = Nueva fórmula: Contrato OK / Con Cobertura (de MANTRA)
        efectividad = get_conversion_asesor_mes(asesor, mes_nombre)
        
        metricas[asesor] = {
            'instalados': instalados,
            'cancelados': cancelados,
            'cumplimiento': cumplimiento,
            'efectividad': efectividad
        }
    
    return metricas

# Cargar datos
def load_data(mes_seleccionado=None):
    # Cargar la hoja LISTA para obtener metas por mes
    df_lista = load_lista_metas()
    
    # Crear diccionario de metas para el mes seleccionado
    metas_dict = {}
    
    if df_lista is not None and not df_lista.empty:
        # Filtrar por el mes seleccionado
        df_mes_metas = df_lista[df_lista['Mes'] == mes_seleccionado].copy()
        
        # Limpiar espacios en blanco de Asesor y convertir Meta a numérico
        df_mes_metas['Asesor'] = df_mes_metas['Asesor'].astype(str).str.strip()
        df_mes_metas['Meta'] = pd.to_numeric(df_mes_metas['Meta'], errors='coerce').fillna(0)
        
        # Crear diccionario {Asesor: Meta} SOLO con los asesores activos en este mes
        for idx, row in df_mes_metas.iterrows():
            metas_dict[row['Asesor']] = int(row['Meta'])
    
    # Si no hay datos para el mes en LISTA, retornar vacío
    if not metas_dict:
        metas_dict = {}
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    # Obtener métricas de DRIVE filtrando por mes (ahora con mes_nombre)
    metricas = calculate_drive_metrics(metas_dict, mes_filtro=mes_num, mes_nombre=mes_seleccionado)
    
    # Construir DataFrame
    empleados = []
    metas = []
    cumplimientos = []
    efectividades = []
    instaladas = []
    canceladas = []
    
    for empleado, meta in metas_dict.items():
        empleados.append(empleado)
        metas.append(meta)
        
        if empleado in metricas:
            cumplimientos.append(metricas[empleado]['cumplimiento'])
            efectividades.append(metricas[empleado]['efectividad'])
            instaladas.append(metricas[empleado]['instalados'])
            canceladas.append(metricas[empleado]['cancelados'])
        else:
            cumplimientos.append(0)
            efectividades.append(0)
            instaladas.append(0)
            canceladas.append(0)
    
    data = {
        'Asesor': empleados,
        'Meta': metas,
        'Instaladas': instaladas,
        'Canceladas': canceladas,
        'Cumplimiento': cumplimientos,
        'Efectividad': efectividades
    }
    
    df = pd.DataFrame(data)
    df['% Meta Alcanzado'] = (df['Cumplimiento'] / 100 * 100).astype(int)
    df['Diferencia'] = df['Cumplimiento'] - 100
    
    # Excluir agentes específicos por mes
    if mes_seleccionado == 'Abril':
        # Excluir ZIM_KATHERINEMM_VTP de Abril
        df = df[df['Asesor'] != 'ZIM_KATHERINEMM_VTP'].reset_index(drop=True)
    
    return df

@st.cache_data(ttl=60)  # Cache de 60 segundos para desarrollo
def load_data_codigo_carga(mes_seleccionado=None):
    """Carga datos agrupados por CODIGO DE CARGA (Agente) para un mes exacto.
    Incluye TODOS los agentes de MANTRA, incluso aquellos sin registros en DRIVE.
    - LEADS vienen de MANTRA (cantidad de registros por Agente)
    - VENTAS = todos los registros del mes (sin importar PAGO o ESTADO)
    - PEND = cantidad de PENDIENTES
    Filtra por columna MES exacto en ambas hojas."""
    df_drive = load_drive_data()
    df_mantra = load_mantra_data()
    
    if df_drive is None or df_drive.empty or df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # ============= LEADS DESDE MANTRA =============
    # Filtrar por MES exacto en MANTRA
    df_mantra_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mantra_mes.empty:
        return pd.DataFrame()
    
    # Limpiar espacios en blanco en Agente y estandarizar (agregar _VTP si no lo tiene)
    df_mantra_mes['Agente'] = df_mantra_mes['Agente'].astype(str).str.strip()
    # Estandarizar: si no termina con _VTP, agregar lo
    df_mantra_mes['Agente'] = df_mantra_mes['Agente'].apply(
        lambda x: x if x.endswith('_VTP') else x + '_VTP'
    )
    
    # Agrupar por Agente y contar LEADS
    leads_dict = df_mantra_mes.groupby('Agente').size().to_dict()
    
    # ============= ESTADÍSTICAS DESDE DRIVE =============
    # Filtrar por MES exacto en DRIVE
    df_drive_mes = df_drive[df_drive['MES'] == mes_seleccionado].copy()
    
    if df_drive_mes.empty:
        df_drive_mes = pd.DataFrame()
    else:
        # Limpiar espacios en blanco en columnas clave
        df_drive_mes['CODIGO DE CARGA'] = df_drive_mes['CODIGO DE CARGA'].astype(str).str.strip()
        df_drive_mes['ESTADO'] = df_drive_mes['ESTADO'].astype(str).str.strip()
        df_drive_mes['PAGO'] = df_drive_mes['PAGO'].astype(str).str.strip()
        df_drive_mes['FECHA'] = pd.to_datetime(df_drive_mes['FECHA'], errors='coerce')
        
        # Nota: No hacemos validación adicional de fecha ya que la columna MES es la fuente de verdad
        # Evitamos doble filtrado que podría excluir registros válidos
    
    # Agrupar por CODIGO DE CARGA y contar estados
    grupos = []
    
    # Obtener todos los agentes únicos de MANTRA (que son los CODIGO DE CARGA)
    agentes_unicos = sorted(leads_dict.keys())
    
    for agente in agentes_unicos:
        # Leads desde MANTRA
        leads = leads_dict.get(agente, 0)
        
        # Con cobertura desde MANTRA (NIVEL 2 = 'Con Cobertura')
        df_agente_mantra = df_mantra_mes[df_mantra_mes['Agente'] == agente]
        df_agente_mantra['NIVEL 2'] = df_agente_mantra['NIVEL 2'].astype(str).str.strip()
        con_cobertura = len(df_agente_mantra[df_agente_mantra['NIVEL 2'] == 'Con Cobertura'])
        
        # Inicializar contadores
        ventas = 0
        pendientes = 0
        
        # Si hay datos en DRIVE, buscar registros del agente
        if not df_drive_mes.empty:
            df_agente = df_drive_mes[df_drive_mes['CODIGO DE CARGA'] == agente]
            
            if not df_agente.empty:
                # VENTAS = registros con PAGO (no vacío)
                ventas = len(df_agente[(df_agente['PAGO'] != '') & (df_agente['PAGO'] != 'nan') & (df_agente['PAGO'].notna())])
                
                # PENDIENTES = registros con ESTADO='PENDIENTE'
                pendientes = len(df_agente[df_agente['ESTADO'] == 'PENDIENTE'])
        
        grupos.append({
            'CODIGO_CARGA': agente,
            'LEADS': leads,
            'CON_COBERTURA': con_cobertura,
            'VENTAS': ventas,
            'PENDIENTES': pendientes
        })
    
    if not grupos:
        return pd.DataFrame()
    
    df_resultado = pd.DataFrame(grupos)
    
    # Calcular % Conversión de Ventas respecto a Leads: (VENTAS / LEADS) * 100
    df_resultado['CONV_VENTAS'] = (df_resultado['VENTAS'] / df_resultado['LEADS'] * 100).round(2)
    
    # Calcular % Conversión de Ventas respecto a Con Cobertura: (VENTAS / CON_COBERTURA) * 100
    # Evitar división por cero
    df_resultado['CONV_VENTAS_COB'] = df_resultado.apply(
        lambda row: round((row['VENTAS'] / row['CON_COBERTURA'] * 100), 2) if row['CON_COBERTURA'] > 0 else 0.0,
        axis=1
    )
    
    # Calcular ventas necesarias para llegar al 10%: (LEADS * 0.10) - VENTAS
    df_resultado['VENTAS_FALTA_10'] = ((df_resultado['LEADS'] * 0.10) - df_resultado['VENTAS']).round(0).astype(int)
    # Si ya alcanzó el 10%, mostrar 0
    df_resultado['VENTAS_FALTA_10'] = df_resultado['VENTAS_FALTA_10'].apply(lambda x: max(0, x))
    
    # Ordenar por VENTAS de mayor a menor
    df_resultado = df_resultado.sort_values('VENTAS', ascending=False).reset_index(drop=True)
    
    # Agregar posición
    df_resultado.insert(0, 'POS', range(1, len(df_resultado) + 1))
    
    # Excluir agentes específicos por mes
    if mes_seleccionado == 'Abril':
        # Excluir ZIM_KATHERINEMM_VTP de Abril
        df_resultado = df_resultado[df_resultado['CODIGO_CARGA'] != 'ZIM_KATHERINEMM_VTP'].reset_index(drop=True)
        # Recalcular posición después de filtrar
        df_resultado['POS'] = range(1, len(df_resultado) + 1)
    
    return df_resultado

# ============= ANÁLISIS DE LEADS CON COBERTURA POR HORA Y FECHA =============

@st.cache_data(ttl=3600)
def get_leads_cobertura_por_hora_fecha(mes_seleccionado="Mayo", asesor_seleccionado="Todos"):
    """Obtiene tabla de Leads con Cobertura (NIVEL 2) por Hora y Fecha desde MANTRA
    Retorna un DataFrame pivote con horas en filas y fechas en columnas
    Si asesor_seleccionado != "Todos", filtra por ese agente/agentes específico(s)
    asesor_seleccionado puede ser string "Todos" o una lista de asesores"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Filtrar por asesor si está especificado
    if asesor_seleccionado != "Todos":
        df_mes['Agente'] = df_mes['Agente'].astype(str).str.strip()
        # Manejar tanto string como lista de asesores
        if isinstance(asesor_seleccionado, list):
            df_mes = df_mes[df_mes['Agente'].isin(asesor_seleccionado)]
        else:
            df_mes = df_mes[df_mes['Agente'] == asesor_seleccionado]
        if df_mes.empty:
            return pd.DataFrame()
    
    # Limpiar espacios en blanco
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Filtrar solo los que tienen "Con Cobertura"
    df_cobertura = df_mes[df_mes['NIVEL 2'] == 'Con Cobertura'].copy()
    
    if df_cobertura.empty:
        return pd.DataFrame()
    
    # Convertir FECHA a datetime y extraer solo la fecha (sin hora)
    df_cobertura['Fecha'] = pd.to_datetime(df_cobertura['Fecha'], errors='coerce')
    df_cobertura['Dia'] = df_cobertura['Fecha'].dt.strftime('%d/%m')
    
    # Convertir HORA a int y filtrar valores válidos (no NaN)
    df_cobertura['Hora'] = df_cobertura['HORA'].dropna().astype(int) if 'HORA' in df_cobertura.columns else None
    
    # Filtrar filas donde HORA no es NaN
    df_cobertura = df_cobertura.dropna(subset=['HORA']).copy()
    df_cobertura['Hora'] = df_cobertura['HORA'].astype(int)
    
    if df_cobertura.empty:
        return pd.DataFrame()
    
    # Agrupar por hora y día
    df_pivot = df_cobertura.groupby(['Hora', 'Dia']).size().reset_index(name='Cantidad')
    
    # Crear tabla pivote (horas en filas, fechas en columnas)
    tabla_pivot = df_pivot.pivot(index='Hora', columns='Dia', values='Cantidad').fillna(0).astype(int)
    
    # Agregar fila de totales por fecha
    totales_por_fecha = tabla_pivot.sum()
    tabla_pivot.loc['TOTAL'] = totales_por_fecha
    
    # Agregar columna de totales por hora
    tabla_pivot['TOTAL'] = tabla_pivot.sum(axis=1)
    
    # Ordenar índice (horas de menor a mayor)
    horas_validas = sorted([h for h in tabla_pivot.index if isinstance(h, (int, float)) and h != 'TOTAL'])
    horas_orden = horas_validas + ['TOTAL']
    tabla_pivot = tabla_pivot.reindex(horas_orden, fill_value=0)
    
    return tabla_pivot

@st.cache_data(ttl=60)
def get_leads_contrato_ok_por_hora_fecha(mes_seleccionado="Mayo", asesor_seleccionado="Todos"):
    """Obtiene tabla de Leads con Contrato OK (NIVEL 3) por Hora y Fecha desde MANTRA
    Retorna un DataFrame pivote con horas en filas y fechas en columnas
    Solo cuenta los que tienen NIVEL 2 = 'Con Cobertura' Y NIVEL 3 = 'Contrato OK'
    asesor_seleccionado puede ser string "Todos" o una lista de asesores"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Filtrar por asesor si está especificado
    if asesor_seleccionado != "Todos":
        df_mes['Agente'] = df_mes['Agente'].astype(str).str.strip()
        # Manejar tanto string como lista de asesores
        if isinstance(asesor_seleccionado, list):
            df_mes = df_mes[df_mes['Agente'].isin(asesor_seleccionado)]
        else:
            df_mes = df_mes[df_mes['Agente'] == asesor_seleccionado]
        if df_mes.empty:
            return pd.DataFrame()
    
    # Limpiar espacios en blanco
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    df_mes['NIVEL 3'] = df_mes['NIVEL 3'].astype(str).str.strip()
    
    # Filtrar solo los que tienen "Con Cobertura" Y "Contrato OK"
    df_contrato_ok = df_mes[
        (df_mes['NIVEL 2'] == 'Con Cobertura') & 
        (df_mes['NIVEL 3'] == 'Contrato OK')
    ].copy()
    
    if df_contrato_ok.empty:
        return pd.DataFrame()
    
    # Convertir FECHA a datetime y extraer solo la fecha (sin hora)
    df_contrato_ok['Fecha'] = pd.to_datetime(df_contrato_ok['Fecha'], errors='coerce')
    df_contrato_ok['Dia'] = df_contrato_ok['Fecha'].dt.strftime('%d/%m')
    
    # Convertir HORA a int y filtrar valores válidos (no NaN)
    df_contrato_ok = df_contrato_ok.dropna(subset=['HORA']).copy()
    df_contrato_ok['Hora'] = df_contrato_ok['HORA'].astype(int)
    
    if df_contrato_ok.empty:
        return pd.DataFrame()
    
    # Agrupar por hora y día
    df_pivot = df_contrato_ok.groupby(['Hora', 'Dia']).size().reset_index(name='Cantidad')
    
    # Crear tabla pivote (horas en filas, fechas en columnas)
    tabla_pivot = df_pivot.pivot(index='Hora', columns='Dia', values='Cantidad').fillna(0).astype(int)
    
    # Agregar fila de totales por fecha
    totales_por_fecha = tabla_pivot.sum()
    tabla_pivot.loc['TOTAL'] = totales_por_fecha
    
    # Agregar columna de totales por hora
    tabla_pivot['TOTAL'] = tabla_pivot.sum(axis=1)
    
    # Ordenar índice (horas de menor a mayor)
    horas_validas = sorted([h for h in tabla_pivot.index if isinstance(h, (int, float)) and h != 'TOTAL'])
    horas_orden = horas_validas + ['TOTAL']
    tabla_pivot = tabla_pivot.reindex(horas_orden, fill_value=0)
    
    return tabla_pivot

@st.cache_data(ttl=60)
def get_detalle_leads_cobertura(mes_seleccionado="Mayo", asesor_seleccionado="Todos", hora_seleccionada="Todos", fecha_seleccionada=None):
    """Obtiene el detalle de quiénes (asesores) se conectaron con cobertura en una hora y fecha específica
    Retorna un DataFrame con los asesores que tuvieron leads en esa combinación
    Si hora_seleccionada es "Todos", no filtra por hora
    asesor_seleccionado puede ser string "Todos" o una lista de asesores"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Filtrar por asesor si está especificado
    if asesor_seleccionado != "Todos":
        df_mes['Agente'] = df_mes['Agente'].astype(str).str.strip()
        # Manejar tanto string como lista de asesores
        if isinstance(asesor_seleccionado, list):
            df_mes = df_mes[df_mes['Agente'].isin(asesor_seleccionado)]
        else:
            df_mes = df_mes[df_mes['Agente'] == asesor_seleccionado]
        if df_mes.empty:
            return pd.DataFrame()
    
    # Limpiar espacios en blanco
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Filtrar solo los que tienen "Con Cobertura"
    df_cobertura = df_mes[df_mes['NIVEL 2'] == 'Con Cobertura'].copy()
    
    if df_cobertura.empty:
        return pd.DataFrame()
    
    # Convertir FECHA a datetime y extraer solo la fecha
    df_cobertura['Fecha'] = pd.to_datetime(df_cobertura['Fecha'], errors='coerce')
    df_cobertura['Dia'] = df_cobertura['Fecha'].dt.strftime('%d/%m')
    
    # Convertir HORA a int y filtrar valores válidos
    df_cobertura = df_cobertura.dropna(subset=['HORA']).copy()
    df_cobertura['Hora'] = df_cobertura['HORA'].astype(int)
    
    # Filtrar por hora si está especificada (y no es "Todos")
    if hora_seleccionada is not None and hora_seleccionada != "Todos":
        df_cobertura = df_cobertura[df_cobertura['Hora'] == hora_seleccionada]
    
    # Filtrar por fecha si está especificada
    if fecha_seleccionada is not None:
        df_cobertura = df_cobertura[df_cobertura['Dia'] == fecha_seleccionada]
    
    if df_cobertura.empty:
        return pd.DataFrame()
    
    # Limpiar columna Agente
    df_cobertura['Agente'] = df_cobertura['Agente'].astype(str).str.strip()
    
    # Retornar solo columnas relevantes: Agente, Hora, Fecha, y otros detalles disponibles
    columnas_disponibles = ['Agente', 'Hora', 'Dia']
    for col in ['Cliente', 'Teléfono', 'Departamento']:
        if col in df_cobertura.columns:
            columnas_disponibles.append(col)
    
    return df_cobertura[columnas_disponibles]

# ============= ANÁLISIS DETALLADO DEL DRIVE =============

@st.cache_data(ttl=3600)
def get_drive_history_by_asesor(asesor, mes_seleccionado="Marzo"):
    """Obtiene historial detallado de transacciones por asesor en el DRIVE"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    # Filtrar por mes y asesor
    df_mes = df_drive[df_drive['MES'] == mes_seleccionado].copy() if 'MES' in df_drive.columns else df_drive
    
    # Limpiar espacios en la columna ASESOR
    df_mes['ASESOR'] = df_mes['ASESOR'].astype(str).str.strip()
    asesor_clean = asesor.strip()
    
    df_asesor = df_mes[df_mes['ASESOR'] == asesor_clean].copy()
    
    if df_asesor.empty:
        return pd.DataFrame()
    
    # Limpiar fechas
    df_asesor['FECHA'] = pd.to_datetime(df_asesor['FECHA'], errors='coerce')
    
    # Ordenar por fecha
    df_asesor = df_asesor.sort_values('FECHA')
    
    # Agregar número secuencial de venta
    df_asesor.insert(0, 'VENTA_NUM', range(1, len(df_asesor) + 1))
    
    return df_asesor

@st.cache_data(ttl=3600)
def get_drive_asesor_kpis(asesor, mes_seleccionado="Marzo"):
    """Calcula KPIs importantes para un asesor en el DRIVE"""
    df_asesor = get_drive_history_by_asesor(asesor, mes_seleccionado)
    
    if df_asesor.empty:
        return {}
    
    # Limpiar columnas
    df_asesor['ESTADO'] = df_asesor['ESTADO'].astype(str).str.strip()
    df_asesor['PAGO'] = df_asesor['PAGO'].astype(str).str.strip()
    df_asesor['FECHA'] = pd.to_datetime(df_asesor['FECHA'], errors='coerce')
    
    total_ventas = len(df_asesor)
    instaladas = len(df_asesor[df_asesor['ESTADO'] == 'INSTALADO'])
    pendientes = len(df_asesor[df_asesor['ESTADO'] == 'PENDIENTE'])
    canceladas = len(df_asesor[df_asesor['ESTADO'] == 'CANCELADO'])
    
    # Tasa de conversión (instaladas / total)
    tasa_conversion = (instaladas / total_ventas * 100) if total_ventas > 0 else 0
    
    # Tasa de cancelación
    tasa_cancelacion = (canceladas / total_ventas * 100) if total_ventas > 0 else 0
    
    # Tiempo promedio entre ventas (en días)
    fechas_validas = df_asesor['FECHA'].dropna()
    if len(fechas_validas) > 1:
        dias_totales = (fechas_validas.max() - fechas_validas.min()).days
        tiempo_promedio = dias_totales / (len(fechas_validas) - 1) if len(fechas_validas) > 1 else 0
    else:
        tiempo_promedio = 0
    
    # Velocidad de venta (ventas por día)
    dias_activos = len(df_asesor['FECHA'].dt.date.unique()) if len(fechas_validas) > 0 else 1
    velocidad_venta = total_ventas / dias_activos if dias_activos > 0 else 0
    
    # Estabilidad (desviación estándar de ventas por día)
    ventas_por_dia = df_asesor.groupby(df_asesor['FECHA'].dt.date).size()
    estabilidad = ventas_por_dia.std() if len(ventas_por_dia) > 1 else 0
    
    return {
        'total_ventas': total_ventas,
        'instaladas': instaladas,
        'pendientes': pendientes,
        'canceladas': canceladas,
        'tasa_conversion': round(tasa_conversion, 1),
        'tasa_cancelacion': round(tasa_cancelacion, 1),
        'tiempo_promedio_dias': round(tiempo_promedio, 1),
        'velocidad_venta': round(velocidad_venta, 2),
        'dias_activos': dias_activos,
        'estabilidad_ventas': round(estabilidad, 2),
        'fecha_primera_venta': fechas_validas.min() if len(fechas_validas) > 0 else None,
        'fecha_ultima_venta': fechas_validas.max() if len(fechas_validas) > 0 else None,
    }

@st.cache_data(ttl=3600)
def get_ventas_mes_pasado(asesor, mes_actual="Marzo"):
    """Obtiene ventas del mes anterior que aún están pendientes"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    # Filtrar por asesor
    df_drive['ASESOR'] = df_drive['ASESOR'].astype(str).str.strip()
    df_asesor = df_drive[df_drive['ASESOR'] == asesor.strip()].copy()
    
    if df_asesor.empty:
        return pd.DataFrame()
    
    # Obtener mes anterior
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    if mes_actual in meses:
        idx_actual = meses.index(mes_actual)
        mes_anterior = meses[idx_actual - 1] if idx_actual > 0 else None
    else:
        mes_anterior = None
    
    if mes_anterior is None:
        return pd.DataFrame()
    
    # Filtrar por mes anterior
    df_anterior = df_asesor[df_asesor['MES'] == mes_anterior].copy()
    
    if df_anterior.empty:
        return pd.DataFrame()
    
    # Limpiar y ordenar
    df_anterior['FECHA'] = pd.to_datetime(df_anterior['FECHA'], errors='coerce')
    df_anterior = df_anterior.sort_values('FECHA')
    
    return df_anterior

@st.cache_data(ttl=3600)
def get_desglose_diario(asesor, mes_seleccionado="Marzo"):
    """Obtiene desglose de ventas por día del mes actual"""
    df_asesor = get_drive_history_by_asesor(asesor, mes_seleccionado)
    
    if df_asesor.empty:
        return pd.DataFrame()
    
    df_asesor['FECHA'] = pd.to_datetime(df_asesor['FECHA'], errors='coerce')
    df_asesor['ESTADO'] = df_asesor['ESTADO'].astype(str).str.strip()
    
    # Agrupar por día y estado
    desglose = df_asesor.groupby([df_asesor['FECHA'].dt.date, 'ESTADO']).size().unstack(fill_value=0)
    desglose['TOTAL'] = desglose.sum(axis=1)
    
    return desglose

@st.cache_data(ttl=3600)
def get_crecimiento_ventas(asesor, mes_seleccionado="Marzo"):
    """Obtiene el crecimiento acumulado de ventas por día con promedio"""
    df_asesor = get_drive_history_by_asesor(asesor, mes_seleccionado)
    
    if df_asesor.empty:
        return pd.DataFrame()
    
    df_asesor['FECHA'] = pd.to_datetime(df_asesor['FECHA'], errors='coerce')
    df_asesor['ESTADO'] = df_asesor['ESTADO'].astype(str).str.strip()
    
    # Contar ventas diarias por tipo
    ventas_diarias = df_asesor.groupby([df_asesor['FECHA'].dt.date, 'ESTADO']).size().unstack(fill_value=0)
    
    # Calcular acumuladas
    crecimiento = pd.DataFrame()
    crecimiento['Fecha'] = ventas_diarias.index
    
    # Usar columnas si existen, sino 0
    instaladas = ventas_diarias['INSTALADO'].values if 'INSTALADO' in ventas_diarias.columns else np.zeros(len(ventas_diarias))
    canceladas = ventas_diarias['CANCELADO'].values if 'CANCELADO' in ventas_diarias.columns else np.zeros(len(ventas_diarias))
    pendientes = ventas_diarias['PENDIENTE'].values if 'PENDIENTE' in ventas_diarias.columns else np.zeros(len(ventas_diarias))
    
    crecimiento['TOTAL'] = instaladas + canceladas + pendientes
    crecimiento['Instaladas'] = instaladas
    crecimiento['Canceladas'] = canceladas
    crecimiento['Pendientes'] = pendientes
    
    # Calcular acumuladas
    crecimiento['Total Acumulado'] = crecimiento['TOTAL'].cumsum()
    crecimiento['Instaladas Acumuladas'] = crecimiento['Instaladas'].cumsum()
    
    # Calcular promedio diario
    promedio_diario = crecimiento['TOTAL'].mean()
    crecimiento['Promedio'] = promedio_diario
    crecimiento['Estado_Promedio'] = crecimiento['TOTAL'].apply(
        lambda x: 'Arriba del Promedio' if x >= promedio_diario else 'Bajo el Promedio'
    )
    
    return crecimiento

@st.cache_data(ttl=3600)
def get_crecimiento_ventas_semanal(asesor, mes_seleccionado="Marzo"):
    """Obtiene el conteo de PAGO (sin importar estado) agrupado por semana
    Rangos de semana personalizados por mes:
    Mayo: Semana 1 (1-3), Semana 2 (4-10), Semana 3 (11-17), Semana 4 (18-24), Semana 5 (25-31)
    Otros meses: Semana 1 (1-5), Semana 2 (6-12), Semana 3 (13-19), Semana 4 (20-26), Semana 5 (27-30)
    NOTA: Las ventas del mes anterior se cuentan en Semana 1"""
    df_asesor = get_drive_history_by_asesor(asesor, mes_seleccionado)
    
    if df_asesor.empty:
        return pd.DataFrame()
    
    df_asesor['FECHA'] = pd.to_datetime(df_asesor['FECHA'], errors='coerce')
    df_asesor['PAGO'] = df_asesor['PAGO'].astype(str).str.strip()
    
    # Filtrar solo registros con PAGO (no vacío)
    df_con_pago = df_asesor[(df_asesor['PAGO'] != '') & (df_asesor['PAGO'] != 'nan') & (df_asesor['PAGO'].notna())].copy()
    
    if df_con_pago.empty:
        return pd.DataFrame()
    
    # Mapeo de meses a números
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num_seleccionado = mes_numeros.get(mes_seleccionado, 5)
    
    # Crear etiqueta descriptiva para cada semana según rangos específicos por mes
    def get_semana_label(row):
        fecha = row['FECHA']
        day = fecha.day
        mes_fecha = fecha.month
        
        # Si la fecha es de un mes diferente (mes anterior), asignarla al inicio (NUM_SEMANA = 0)
        if mes_fecha != mes_num_seleccionado:
            return (0, "Semana 1 (Mes Anterior)")
        
        # Rangos específicos para Mayo
        if mes_seleccionado == "Mayo":
            if day <= 3:
                return (1, "Semana 1 (1-3)")
            elif day <= 10:
                return (2, "Semana 2 (4-10)")
            elif day <= 17:
                return (3, "Semana 3 (11-17)")
            elif day <= 24:
                return (4, "Semana 4 (18-24)")
            else:
                return (5, "Semana 5 (25-31)")
        else:
            # Rangos por defecto para otros meses
            if day <= 5:
                return (1, "Semana 1 (1-5)")
            elif day <= 12:
                return (2, "Semana 2 (6-12)")
            elif day <= 19:
                return (3, "Semana 3 (13-19)")
            elif day <= 26:
                return (4, "Semana 4 (20-26)")
            else:
                return (5, "Semana 5 (27-30)")
    
    df_con_pago[['NUM_SEMANA', 'LABEL_SEMANA']] = df_con_pago.apply(lambda row: pd.Series(get_semana_label(row)), axis=1)
    
    # Contar total de PAGO por semana
    pagos_semanales = df_con_pago.groupby(['NUM_SEMANA', 'LABEL_SEMANA']).size().reset_index(name='TOTAL')
    pagos_semanales = pagos_semanales.sort_values('NUM_SEMANA')
    
    crecimiento_semanal = pd.DataFrame()
    crecimiento_semanal['Semana'] = pagos_semanales['LABEL_SEMANA'].values
    crecimiento_semanal['TOTAL'] = pagos_semanales['TOTAL'].values
    
    # Calcular promedio semanal
    promedio_semanal = crecimiento_semanal['TOTAL'].mean()
    crecimiento_semanal['Promedio'] = promedio_semanal
    crecimiento_semanal['Estado_Promedio'] = crecimiento_semanal['TOTAL'].apply(
        lambda x: 'Arriba del Promedio' if x >= promedio_semanal else 'Bajo el Promedio'
    )
    
    return crecimiento_semanal

@st.cache_data(ttl=3600)
def get_drive_tendencias(asesor, mes_seleccionado="Marzo"):
    """Analiza tendencias de ventas semana a semana"""
    df_asesor = get_drive_history_by_asesor(asesor, mes_seleccionado)
    
    if df_asesor.empty:
        return pd.DataFrame()
    
    df_asesor['FECHA'] = pd.to_datetime(df_asesor['FECHA'], errors='coerce')
    df_asesor['ESTADO'] = df_asesor['ESTADO'].astype(str).str.strip()
    df_asesor['SEMANA'] = df_asesor['FECHA'].dt.isocalendar().week
    
    # Contar por semana y estado
    tendencias = df_asesor.groupby(['SEMANA', 'ESTADO']).size().unstack(fill_value=0)
    
    return tendencias

def get_recomendaciones_asesor(asesor, kpis, mes_seleccionado="Marzo"):
    """Genera recomendaciones personalizadas basadas en el análisis"""
    recomendaciones = []
    
    if kpis.get('tasa_conversion', 0) < 50:
        recomendaciones.append({
            'tipo': 'crítica',
            'título': '⚠️ Tasa de Conversión Baja',
            'descripción': f"Su tasa de conversión es {kpis.get('tasa_conversion', 0)}%. Objetivo: aumentar a 80%+",
            'acción': 'Enfoque en filtrar mejor los leads antes de hacer seguimiento. Calidad > Cantidad'
        })
    
    if kpis.get('tasa_cancelacion', 0) > 30:
        recomendaciones.append({
            'tipo': 'crítica',
            'título': '⚠️ Tasa de Cancelación Alta',
            'descripción': f"El {kpis.get('tasa_cancelacion', 0)}% de sus ventas se cancelan",
            'acción': 'Investigar causas de cancelación. ¿Falta de seguimiento? ¿Expectativas no cumplidas?'
        })
    
    if kpis.get('velocidad_venta', 0) < 1:
        recomendaciones.append({
            'tipo': 'alta',
            'título': '📈 Aumentar Velocidad de Venta',
            'descripción': f"Actualmente {kpis.get('velocidad_venta', 0)} ventas/día. Potencial: 2+ ventas/día",
            'acción': 'Dedique más tiempo a llamadas y seguimiento. Cree rutinas diarias'
        })
    
    if kpis.get('estabilidad_ventas', 0) > 2:
        recomendaciones.append({
            'tipo': 'media',
            'título': '📊 Variabilidad en Ventas',
            'descripción': f"Sus ventas varían mucho por día (desv. std: {kpis.get('estabilidad_ventas', 0)})",
            'acción': 'Implemente una estrategia consistente diaria. Objetivos claros por día'
        })
    
    if kpis.get('pendientes', 0) > 5:
        recomendaciones.append({
            'tipo': 'media',
            'título': '⏳ Muchas Ventas Pendientes',
            'descripción': f"Tiene {kpis.get('pendientes', 0)} ventas en estado PENDIENTE",
            'acción': 'Prioricel el seguimiento de pendientes para convertirlas en instaladas'
        })
    
    if len(recomendaciones) == 0:
        recomendaciones.append({
            'tipo': 'éxito',
            'título': '✅ Desempeño Sólido',
            'descripción': "Sus métricas se ven bien. Continúe con esta tendencia",
            'acción': 'Mantenga la consistencia y búsque formas de optimizar aún más'
        })
    
    return recomendaciones

# ============= ANÁLISIS DE METAS =============

@st.cache_data(ttl=3600)
def load_metas_data():
    """Carga los datos de metas de la hoja METAS del archivo REPORTE FTTH.xlsx"""
    excel_path = os.path.join(os.path.dirname(__file__), 'REPORTE FTTH.xlsx')
    
    try:
        df_metas = pd.read_excel(excel_path, sheet_name='METAS')
        # Limpiar columnas
        df_metas.columns = df_metas.columns.str.strip()
        # Mapeo de nombres de asesores para evitar inconsistencias
        df_metas['ASESOR'] = df_metas['ASESOR'].astype(str).str.strip()
        return df_metas
    except Exception as e:
        return None

def get_semana_actual(mes_seleccionado="Mayo"):
    """Determina la semana actual basada en la fecha de hoy y los rangos de semanas del mes"""
    hoy = datetime.now().date()
    day = hoy.day
    
    # Rangos específicos para Mayo
    if mes_seleccionado == "Mayo":
        if day <= 3:
            return 1
        elif day <= 10:
            return 2
        elif day <= 17:
            return 3
        elif day <= 24:
            return 4
        else:
            return 5
    else:
        # Rangos por defecto para otros meses
        if day <= 5:
            return 1
        elif day <= 12:
            return 2
        elif day <= 19:
            return 3
        elif day <= 26:
            return 4
        else:
            return 5

@st.cache_data(ttl=60)
def get_cumplimiento_metas_analisis(asesor_seleccionado, mes_seleccionado="Mayo"):
    """Calcula el análisis de cumplimiento de metas: compara ventas vs metas diarias, semanales y mensuales"""
    df_metas = load_metas_data()
    df_drive = load_drive_data()
    
    if df_metas is None or df_drive is None:
        return None
    
    # Obtener metas del asesor
    if asesor_seleccionado == "Todos":
        # Si es Todos, retornar análisis de todos los asesores
        return None  # Se manejará de forma diferente
    
    # Buscar el asesor en la hoja METAS
    metas_asesor = df_metas[df_metas['ASESOR'] == asesor_seleccionado]
    
    if metas_asesor.empty:
        return None
    
    # Extraer metas
    meta_diaria = metas_asesor['Meta Diaria'].values[0]
    meta_semanal = metas_asesor['Meta semanal'].values[0]
    meta_mensual = metas_asesor['Meta Mensual'].values[0]
    
    # Obtener ventas del asesor en el mes
    df_asesor_drive = get_drive_history_by_asesor(asesor_seleccionado, mes_seleccionado)
    
    if df_asesor_drive.empty:
        ventas_diarias = {}
        ventas_semanales = {}
        ventas_totales_todos = 0
        progreso_semana_actual = 0
    else:
        df_asesor_drive['FECHA'] = pd.to_datetime(df_asesor_drive['FECHA'], errors='coerce')
        df_asesor_drive['ESTADO'] = df_asesor_drive['ESTADO'].astype(str).str.strip()
        
        # Total de TODAS las ventas brutas (INSTALADO + PENDIENTE + CANCELADO)
        ventas_totales_todos = len(df_asesor_drive)
        
        # Ventas por día (TODAS las ventas brutas)
        ventas_diarias_counts = df_asesor_drive.groupby(df_asesor_drive['FECHA'].dt.date).size()
        ventas_diarias = ventas_diarias_counts.to_dict()
        
        # Ventas por semana (TODAS las ventas brutas)
        df_asesor_drive['SEMANA'] = df_asesor_drive['FECHA'].dt.isocalendar().week
        ventas_semanales_counts = df_asesor_drive.groupby('SEMANA').size()
        ventas_semanales = ventas_semanales_counts.to_dict()
        
        # Calcular progreso de la semana actual
        semana_actual = get_semana_actual(mes_seleccionado)
        progreso_semana_actual = 0
        
        # Mapeo de meses a números
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num_seleccionado = mes_numeros.get(mes_seleccionado, 5)
        
        # Obtener rango de días de la semana actual
        if mes_seleccionado == "Mayo":
            if semana_actual == 1:
                rango_inicio, rango_fin = 1, 3
            elif semana_actual == 2:
                rango_inicio, rango_fin = 4, 10
            elif semana_actual == 3:
                rango_inicio, rango_fin = 11, 17
            elif semana_actual == 4:
                rango_inicio, rango_fin = 18, 24
            else:
                rango_inicio, rango_fin = 25, 31
        else:
            if semana_actual == 1:
                rango_inicio, rango_fin = 1, 5
            elif semana_actual == 2:
                rango_inicio, rango_fin = 6, 12
            elif semana_actual == 3:
                rango_inicio, rango_fin = 13, 19
            elif semana_actual == 4:
                rango_inicio, rango_fin = 20, 26
            else:
                rango_inicio, rango_fin = 27, 31
        
        # Filtrar ventas de la semana actual
        df_semana_actual = df_asesor_drive[
            (df_asesor_drive['FECHA'].dt.day >= rango_inicio) &
            (df_asesor_drive['FECHA'].dt.day <= rango_fin) &
            (df_asesor_drive['FECHA'].dt.month == mes_num_seleccionado)
        ]
        progreso_semana_actual = len(df_semana_actual)
    
    # Calcular cumplimiento usando TODAS las ventas brutas
    promedio_diario = ventas_totales_todos / max(1, len(ventas_diarias)) if ventas_diarias else 0
    
    analisis = {
        'meta_diaria': meta_diaria,
        'meta_semanal': meta_semanal,
        'meta_mensual': meta_mensual,
        'ventas_diarias': ventas_diarias,
        'ventas_semanales': ventas_semanales,
        'ventas_totales_todos': ventas_totales_todos,
        'promedio_diario': promedio_diario,
        'cumplimiento_mensual': (ventas_totales_todos / max(1, meta_mensual) * 100),
        'cumplimiento_diario_prom': (promedio_diario / max(1, meta_diaria) * 100),
        'dias_trabajados': len(ventas_diarias),
        'semanas_activas': len(ventas_semanales),
        'progreso_semana_actual': progreso_semana_actual,
        'cumplimiento_semanal_actual': (progreso_semana_actual / max(1, meta_semanal) * 100)
    }
    
    return analisis

@st.cache_data(ttl=60)
def get_comparativa_metas_todos(mes_seleccionado="Mayo"):
    """Obtiene comparativa de todos los asesores: metas vs ventas reales"""
    df_metas = load_metas_data()
    df_drive = load_drive_data()
    
    if df_metas is None or df_drive is None:
        return pd.DataFrame()
    
    resultados = []
    
    for idx, row in df_metas.iterrows():
        asesor = row['ASESOR']
        meta_mensual = row['Meta Mensual']
        
        # Obtener ventas del asesor
        df_asesor_drive = get_drive_history_by_asesor(asesor, mes_seleccionado)
        
        if not df_asesor_drive.empty:
            # Contar TODAS las ventas brutas (INSTALADO + PENDIENTE + CANCELADO)
            ventas_reales = len(df_asesor_drive)
        else:
            ventas_reales = 0
        
        cumplimiento = (ventas_reales / max(1, meta_mensual) * 100) if meta_mensual > 0 else 0
        
        # Determinar estado
        if cumplimiento >= 100:
            estado = "✅ Cumpliendo"
            color = "#10b981"
        elif cumplimiento >= 80:
            estado = "🟡 Acercándose"
            color = "#f59e0b"
        elif cumplimiento >= 50:
            estado = "🔴 Retrasado"
            color = "#f97316"
        else:
            estado = "❌ Crítico"
            color = "#ef4444"
        
        # Brecha
        brecha = ventas_reales - meta_mensual
        
        resultados.append({
            'Asesor': asesor,
            'Meta': meta_mensual,
            'Ventas': ventas_reales,
            'Cumplimiento': cumplimiento,
            'Estado': estado,
            'Color': color,
            'Brecha': brecha
        })
    
    df_resultados = pd.DataFrame(resultados)
    return df_resultados

@st.cache_data(ttl=60)
def get_oportunidades_mejora(mes_seleccionado="Mayo"):
    """Identifica oportunidades de mejora basadas en cumplimiento de metas"""
    df_comparativa = get_comparativa_metas_todos(mes_seleccionado)
    
    if df_comparativa.empty:
        return {}
    
    oportunidades = {
        'asesores_bajo_cumplimiento': df_comparativa[df_comparativa['Cumplimiento'] < 50].sort_values('Cumplimiento'),
        'asesores_excelentes': df_comparativa[df_comparativa['Cumplimiento'] >= 100].sort_values('Cumplimiento', ascending=False),
        'promedio_equipo': df_comparativa['Cumplimiento'].mean(),
        'brecha_total': df_comparativa['Brecha'].sum(),
        'total_asesores': len(df_comparativa),
        'asesores_en_riesgo': len(df_comparativa[df_comparativa['Cumplimiento'] < 50])
    }
    
    return oportunidades

# Estilos mejorados con tema moderno y premium - CACHEADO
@st.cache_data(ttl=86400)  # Cache por 24 horas
def apply_custom_css():
    """Aplica estilos CSS personalizados una sola vez"""
    st.markdown("""
<style>
    :root {
        --primary-color: #0066cc;
        --secondary-color: #00d4ff;
        --success-color: #10b981;
        --warning-color: #f59e0b;
        --danger-color: #ef4444;
        --dark-bg: #0f172a;
        --light-bg: #f8fafc;
        --card-bg: #ffffff;
        --text-primary: #1e293b;
        --text-secondary: #64748b;
        --border-color: #e2e8f0;
    }

    * {
        font-family: 'Segoe UI', 'Helvetica Neue', sans-serif;
        margin: 0;
        padding: 0;
    }

    body {
        background-color: #f0f4f8;
    }

    /* HEADER PRINCIPAL */
    .header-container {
        background: linear-gradient(135deg, #0066cc 0%, #00d4ff 100%);
        padding: 40px 30px;
        border-radius: 15px;
        color: white;
        box-shadow: 0 10px 40px rgba(0, 102, 204, 0.2);
        margin-bottom: 30px;
        position: relative;
        overflow: hidden;
    }

    .header-container::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -10%;
        width: 300px;
        height: 300px;
        background: rgba(255, 255, 255, 0.1);
        border-radius: 50%;
        z-index: 0;
    }

    .header-content {
        position: relative;
        z-index: 1;
    }

    .header-title {
        font-size: 2.8em;
        font-weight: 800;
        margin-bottom: 10px;
        letter-spacing: -0.5px;
        text-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
    }

    .header-subtitle {
        font-size: 1.3em;
        font-weight: 400;
        opacity: 0.95;
        margin-top: 5px;
    }

    /* KPI CARDS */
    .kpi-container {
        display: grid;
        grid-template-columns: repeat(6, 1fr);
        gap: 12px;
        margin-bottom: 30px;
    }

    .kpi-card {
        background: white;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.06);
        border: 1px solid #e2e8f0;
        transition: all 0.3s ease;
        position: relative;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        min-height: 160px;
    }

    .kpi-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 12px 24px rgba(0, 102, 204, 0.15);
        border-color: #0066cc;
    }

    .kpi-icon {
        font-size: 1.6em;
        margin-bottom: 8px;
    }

    .kpi-value {
        font-size: 2.2em;
        font-weight: 800;
        color: #0066cc;
        margin: 8px 0;
        font-family: 'Courier New', monospace;
        line-height: 1.2;
    }

    .kpi-label {
        font-size: 0.75em;
        color: #64748b;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        line-height: 1.3;
        word-wrap: break-word;
    }

    /* SECCIÓN TÍTULOS */
    .section-title {
        font-size: 1.6em;
        font-weight: 700;
        color: #1e293b;
        margin: 30px 0 20px 0;
        padding-bottom: 12px;
        border-bottom: 3px solid #0066cc;
        display: inline-block;
    }

    /* TABLA META */
    .meta-tabla {
        background: white;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        border: 1px solid #e2e8f0;
    }

    .meta-tabla table {
        width: 100%;
        border-collapse: collapse;
    }

    .meta-tabla thead {
        background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);
        color: white;
        font-weight: 700;
    }

    .meta-tabla th {
        padding: 8px 6px;
        text-align: left;
        font-size: 0.8em;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .meta-tabla td {
        padding: 5px 6px;
        border-bottom: 1px solid #f1f5f9;
        font-size: 0.8em;
    }

    .meta-tabla tbody tr {
        transition: background-color 0.2s ease;
    }

    .meta-tabla tbody tr:hover {
        background-color: #f8fafc;
    }

    .meta-tabla tbody tr:nth-child(odd) {
        background-color: #f9fafc;
    }

    .meta-valor {
        font-weight: 700;
        text-align: center;
        border-radius: 6px;
        padding: 2px 6px;
        display: inline-block;
        min-width: 32px;
        font-size: 0.8em;
        background: linear-gradient(135deg, #f59e0b 0%, #f97316 100%);
        color: white;
    }

    .meta-total {
        background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);
        color: white;
        font-weight: 800 !important;
    }

    .meta-total-row {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        font-weight: 700;
    }

    /* CARDS DE GRÁFICOS */
    .chart-container {
        background: white;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        border: 1px solid #e2e8f0;
    }

    .chart-title {
        font-size: 1.2em;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 15px;
        display: flex;
        align-items: center;
        gap: 10px;
    }

    /* TABLA RESUMEN */
    .resumen-tabla {
        background: white;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        border: 1px solid #e2e8f0;
    }

    .resumen-tabla table {
        width: 100%;
        border-collapse: collapse;
    }

    .resumen-tabla thead {
        background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);
        color: white;
    }

    .resumen-tabla th {
        padding: 10px 6px;
        text-align: center;
        font-size: 0.7em;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }

    .resumen-tabla td {
        padding: 8px 6px;
        border-bottom: 1px solid #f1f5f9;
        text-align: center;
        font-weight: 600;
        font-size: 0.85em;
    }

    .resumen-tabla tbody tr:hover {
        background-color: #f8fafc;
    }

    /* INDICADORES DE ESTADO */
    .status-excellent {
        color: #10b981;
        background: #dcfce7;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 700;
    }

    .status-good {
        color: #f59e0b;
        background: #fef3c7;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 700;
    }

    .status-poor {
        color: #ef4444;
        background: #fee2e2;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 700;
    }

    /* BOTONES */
    .filter-button {
        padding: 10px 16px;
        border: 2px solid #e2e8f0;
        border-radius: 8px;
        background: white;
        color: #1e293b;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.2s ease;
    }

    .filter-button:hover {
        border-color: #0066cc;
        color: #0066cc;
    }

    /* DIVIDER */
    .section-divider {
        margin: 40px 0;
        border: none;
        border-top: 2px solid #e2e8f0;
    }

    /* FOOTER */
    .footer-container {
        text-align: center;
        color: #94a3b8;
        margin-top: 50px;
        padding: 30px;
        font-size: 0.9em;
    }

    .footer-container p {
        margin: 5px 0;
    }

    /* STREAMLIT CUSTOMIZATION */
    [data-testid="stDataFrame"] {
        border-radius: 12px !important;
        overflow: hidden !important;
    }

    /* SELECTBOX STYLING */
    .stSelectbox {
        border-radius: 8px !important;
    }

    /* TABS */
    [data-testid="stTabs"] {
        margin-top: 20px;
    }

    /* SCROLLBAR */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }

    ::-webkit-scrollbar-track {
        background: #f1f5f9;
    }

    ::-webkit-scrollbar-thumb {
        background: #cbd5e1;
        border-radius: 10px;
    }

    ::-webkit-scrollbar-thumb:hover {
        background: #94a3b8;
    }

    @media (max-width: 1400px) {
        .kpi-container {
            grid-template-columns: repeat(3, 1fr);
        }
    }

    @media (max-width: 768px) {
        .kpi-container {
            grid-template-columns: repeat(2, 1fr);
        }

        .header-title {
            font-size: 2em;
        }
    }
</style>
""", unsafe_allow_html=True)

# Aplicar CSS cache al inicio
apply_custom_css()

# Filtros mejorados con layout dinámico
st.markdown("### ⚙️ Filtros y Opciones")
col_filtros = st.columns(3, gap="medium")

with col_filtros[0]:
    mes = st.selectbox("📅 Selecciona Mes", ["Noviembre", "Diciembre", "Enero", "Febrero", "Marzo", "Abril", "Mayo"], index=6)

# Mapeo de meses a años
mes_año_map = {
    "Noviembre": "Noviembre 2025",
    "Diciembre": "Diciembre 2025",
    "Enero": "Enero 2026",
    "Febrero": "Febrero 2026",
    "Marzo": "Marzo 2026",
    "Abril": "Abril 2026",
    "Mayo": "Mayo 2026"
}

# Header mejorado - Dinámico
st.markdown(f"""
<div class="header-container">
    <div class="header-content">
        <div class="header-title">🌐 WORLDTEL</div>
        <div class="header-subtitle">Dashboard de Cumplimiento Mensual - {mes_año_map[mes]}</div>
    </div>
    <div style="position: absolute; right: 250px; top: 50%; transform: translateY(-50%); color: white; font-size: 3.8em; font-weight: 800; letter-spacing: -0.5px;">BITEL - FTTH</div>
</div>
""", unsafe_allow_html=True)

# Cargar datos con el mes seleccionado
df = load_data(mes)

with col_filtros[1]:
    opciones_asesores = ["Todos"] + sorted(df['Asesor'].unique())
    asesor_seleccionado = st.selectbox("👤 Filtrar por Asesor", opciones_asesores)

vista = "Completa"

# Obtener valores del Excel basado en el mes seleccionado
total_leads_excel, _ = get_total_leads_and_conversion(mes)

# Obtener Ventas Del Mes desde DRIVE (SOLO transacciones con FECHA en el mes, sin incluir antiguas)
total_conversion_excel = get_ventas_del_mes_por_fecha(mes)

# KPI Cards mejorados - Datos del asesor seleccionado o totales
def get_cumplimiento_total_mes(mes_nombre):
    """Calcula el cumplimiento total del mes: (Ventas Generales / Meta Global) * 100
    Donde Meta Global = suma de todas las metas individuales"""
    df_drive = load_drive_data()
    df_lista = load_lista_metas()
    
    if df_drive is None or df_drive.empty or df_lista is None or df_lista.empty:
        return 0
    
    try:
        # Obtener meta global: suma de todas las metas del mes
        df_mes_lista = df_lista[df_lista['Mes'] == mes_nombre]
        meta_global = df_mes_lista['Meta'].sum()
        
        if meta_global == 0:
            return 0
        
        # Obtener Ventas Generales del mes
        ventas_generales = get_ventas_generales_mes(mes_nombre)
        
        # Calcular cumplimiento: (Ventas Generales / Meta Global) * 100
        cumplimiento_total = round((ventas_generales / meta_global * 100))
        
        return cumplimiento_total
    except Exception as e:
        return 0

def get_efectividad_mes(mes_nombre):
    """Calcula la efectividad para un mes: INSTALADAS/(INSTALADAS+CANCELADAS)
    Donde INSTALADAS = INSTALADO (sin PENDIENTE)"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    try:
        # Mapear nombre del mes a número
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_nombre)
        
        if mes_num is None:
            return 0
        
        # Filtrar por MES en lugar de FECHA
        if 'MES' in df_drive.columns:
            df_filtrado = df_drive[df_drive['MES'] == mes_nombre]
        else:
            # Fallback a FECHA si MES no existe
            df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
            # Para Noviembre, incluir Octubre + Noviembre
            if mes_num == 11:
                df_filtrado = df_drive[
                    ((df_drive['FECHA'].dt.month == 10) | (df_drive['FECHA'].dt.month == 11))
                ]
            else:
                # Para otros meses, solo ese mes
                df_filtrado = df_drive[
                    (df_drive['FECHA'].dt.month == mes_num)
                ]
        
        # Contar instaladas con regla (solo INSTALADO)
        instaladas = count_instaladas_con_regla(df_filtrado, mes_num, mes_num == 11, mes_nombre)
        canceladas = len(df_filtrado[df_filtrado['ESTADO'] == 'CANCELADO'])
        
        # Calcular efectividad
        total_transacciones = instaladas + canceladas
        if total_transacciones > 0:
            efectividad = round((instaladas / total_transacciones * 100))
        else:
            efectividad = 0
        
        return efectividad
    except Exception as e:
        return 0
        
        # Calcular efectividad
        total_transacciones = instaladas + canceladas
        if total_transacciones > 0:
            efectividad = round((instaladas / total_transacciones * 100))
        else:
            efectividad = 0
        
        return efectividad
    except Exception as e:
        return 0

def get_ventas_mes(mes_nombre):
    """Obtiene el total de instaladas para un mes específico del DRIVE
    Donde INSTALADAS = Solo INSTALADO (no incluye PENDIENTE)
    Filtra por columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    try:
        # Mapear nombre del mes a número
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_nombre)
        
        if mes_num is None:
            return 0
        
        # Convertir FECHA a datetime
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        
        # Para Noviembre, sumar Octubre + Noviembre
        es_noviembre = mes_num == 11
        total = count_instaladas_con_regla(df_drive, mes_num, es_noviembre, mes_nombre)
        return total
    except Exception as e:
        return 0

st.markdown("")  # Espaciador

col1, col2, col3, col4, col5, col6, col7 = st.columns(7, gap="small")

if asesor_seleccionado == "Todos":
    # Preparar datos según vista
    df_vista = df[['Asesor', 'Cumplimiento']].copy()
    
    if vista == "Top 5":
        df_vista = df_vista.nlargest(5, 'Cumplimiento')
    elif vista == "Últimos 5":
        df_vista = df_vista.nsmallest(5, 'Cumplimiento')
    
    # Obtener asesores en la vista
    asesores_vista = df_vista['Asesor'].tolist()
    
    # Obtener ventas totales, efectividad y cumplimiento total del mes actual desde DRIVE
    # SIN filtrar por asesores - mostrar TOTALES de TODOS
    df_drive_filtrado = load_drive_data()
    
    if df_drive_filtrado is not None and not df_drive_filtrado.empty:
        # NO filtrar por asesores - calcular para TODOS
        # df_drive_filtrado = df_drive_filtrado[df_drive_filtrado['ASESOR'].isin(asesores_vista)]
        
        # Calcular métricas para esta vista
        df_drive_filtrado['FECHA'] = pd.to_datetime(df_drive_filtrado['FECHA'], errors='coerce')
        
        # Determinar número de mes
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes, None)
        
        # Filtrar por columna MES (en lugar de por FECHA)
        if 'MES' in df_drive_filtrado.columns:
            df_mes_filtrado = df_drive_filtrado[df_drive_filtrado['MES'] == mes]
        else:
            # Fallback a FECHA si MES no existe
            df_drive_filtrado['FECHA'] = pd.to_datetime(df_drive_filtrado['FECHA'], errors='coerce')
            if mes_num == 11:
                df_mes_filtrado = df_drive_filtrado[
                    ((df_drive_filtrado['FECHA'].dt.month == 10) | (df_drive_filtrado['FECHA'].dt.month == 11))
                ]
            else:
                df_mes_filtrado = df_drive_filtrado[df_drive_filtrado['FECHA'].dt.month == mes_num]
        
        # Ventas (instaladas) - aplicando regla: Solo INSTALADO
        ventas_total = count_instaladas_con_regla(df_mes_filtrado, mes_num, mes_num == 11, mes)
        
        # Efectividad - Nueva fórmula: Contrato OK / Con Cobertura (de MANTRA)
        efectividad_mes = get_conversion_mantra_mes(mes)
        
        # Cumplimiento - NUEVA FÓRMULA: VENTAS DEL MES / 735
        # Fórmula: (Ventas Del Mes / 735) * 100
        cumplimiento_total = round((total_conversion_excel / 735 * 100)) if 735 > 0 else 0
        
        # Ventas generales (total de todas las transacciones)
        ventas_generales = get_ventas_generales_mes(mes)
    else:
        ventas_total = 0
        efectividad_mes = 0
        cumplimiento_total = 0
        ventas_generales = 0
    
    kpis = [
        (f"{total_leads_excel:,}", "📋 Leads", col1),
        (str(get_con_cobertura_count(mes)), "🌐 Con Cobertura", col2),
        (f"{total_conversion_excel}", "✅ Ventas Del Mes", col3),
        (f"{round((total_conversion_excel / total_leads_excel * 100))}%", "📊 % Conversión Total", col4),
        (str(ventas_generales), "📈 Ventas Generales Del Mes", col5),
        (f"{efectividad_mes}%", "⭐ Conversión de Ventas", col6),
        (f"{cumplimiento_total}%", "🎯 Cumplimiento", col7),
    ]
else:
    asesor_data = df[df['Asesor'] == asesor_seleccionado].iloc[0]
    cumpl_val = int(asesor_data['Cumplimiento'])
    efect_val = int(asesor_data['Efectividad'])
    instaladas_asesor = int(asesor_data['Instaladas'])
    leads_asesor = get_leads_asesor_mes(asesor_seleccionado, mes)
    con_cobertura_asesor = get_con_cobertura_asesor_mes(asesor_seleccionado, mes)
    
    kpis = [
        (str(leads_asesor), "📋 Total Leads", col1),
        (str(con_cobertura_asesor), "🌐 Con Cobertura", col2),
        (str(int(asesor_data['Meta'])), "🏆 Meta", col3),
        (f"{cumpl_val}%", "✅ Cumplimiento", col4),
        (f"{efect_val}%", "⭐ Conv. Ventas", col5),
        (str(instaladas_asesor), "💰 Instaladas", col6),
        ("🟢 Excelente" if cumpl_val >= 70 else "🟡 Bueno" if cumpl_val >= 50 else "🔴 Bajo", "📈 Estado", col7),
    ]

for valor, label, col in kpis:
    with col:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-label">{label.split(' ', 1)[1]}</div>
            <div class="kpi-value">{valor}</div>
        </div>
        """, unsafe_allow_html=True)

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Sección principal con 3 columnas mejorada - RESPONSIVO
st.markdown("### 📑 Análisis de Desempeño por Agente")
col1, col2, col3 = st.columns([0.8, 1.6, 1.6], gap="medium")

# Columna 1: Meta Mensual
with col1:
    st.markdown('<div class="chart-title">📈 Meta Mensual</div>', unsafe_allow_html=True)
    tabla_meta = df[['Asesor', 'Meta']].copy()
    tabla_meta = tabla_meta.sort_values('Meta', ascending=False).reset_index(drop=True)
    tabla_meta.index = tabla_meta.index + 1
    
    # Crear HTML para la tabla personalizada
    html_tabla = '<div class="meta-tabla" style="width: auto; max-width: none;"><table><thead><tr><th>Pos</th><th>Asesor</th><th style="text-align: center;">Meta</th></tr></thead><tbody>'
    
    for idx, row in tabla_meta.iterrows():
        asesor = row['Asesor']
        meta = int(row['Meta'])
        html_tabla += f'<tr><td style="font-weight: 700; text-align: center; color: #0066cc;">#{idx}</td><td style="font-weight: 600;">{asesor}</td><td style="text-align: center;"><div class="meta-valor">{meta}</div></td></tr>'
    
    # Agregar fila de totales
    total_meta = int(tabla_meta['Meta'].sum())
    html_tabla += f'<tr style="background-color: #e0e7ff; font-weight: 700; border-top: 2px solid #0066cc;"><td style="text-align: center; color: #0066cc;">∑</td><td style="font-weight: 700; color: #0066cc;">TOTAL</td><td style="text-align: center; font-weight: 700; color: #0066cc;"><div class="meta-valor" style="background-color: #0066cc; color: white;">{total_meta}</div></td></tr>'
    
    html_tabla += '</tbody></table></div>'
    
    st.markdown(html_tabla, unsafe_allow_html=True)

# Columna 2: Cumplimiento por Agente
with col2:
    st.markdown('<div class="chart-title">🎯 Cumplimiento por Agente (%)</div>', unsafe_allow_html=True)
    df_sorted = df.sort_values('Cumplimiento', ascending=True)
    
    # Crear colores degradados basados en cumplimiento
    def get_color(val):
        if val >= 100:
            return '#10b981'  # Verde
        elif val >= 75:
            return '#f59e0b'  # Naranja
        elif val >= 50:
            return '#f97316'  # Naranja oscuro
        else:
            return '#ef4444'  # Rojo
    
    colors = [get_color(x) for x in df_sorted['Cumplimiento']]
    
    fig_cumpl = go.Figure()
    fig_cumpl.add_trace(go.Bar(
        y=df_sorted['Asesor'],
        x=df_sorted['Cumplimiento'],
        orientation='h',
        marker=dict(
            color=colors,
            line=dict(color='white', width=2),
            opacity=0.85
        ),
        text=df_sorted['Cumplimiento'].apply(lambda x: f'{x}%'),
        textposition='outside',
        textfont=dict(size=13, color='#1e293b', family='Arial', weight='bold'),
        hovertemplate='<b>%{y}</b><br><b>Cumplimiento:</b> <b>%{x}%</b><extra></extra>',
        name=''
    ))
    fig_cumpl.update_layout(
        height=580,
        margin=dict(l=160, r=50, t=20, b=20),
        showlegend=False,
        xaxis_title="",
        xaxis=dict(
            gridcolor='rgba(0,0,0,0.05)',
            showgrid=True,
            zeroline=False,
            tickfont=dict(size=11, color='#64748b'),
            range=[0, 130],
            ticksuffix='%'
        ),
        yaxis=dict(
            tickfont=dict(size=9, color='#1e293b'),
            automargin=True
        ),
        plot_bgcolor='rgba(248, 250, 252, 0.5)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=11, family='Arial', color='#1e293b'),
        hovermode='closest'
    )
    st.plotly_chart(fig_cumpl, use_container_width=True, config={'displayModeBar': False})

# Columna 3: Conversión de Ventas por Agente
with col3:
    st.markdown('<div class="chart-title">⭐ Conversión de Ventas por Agente (%)</div>', unsafe_allow_html=True)
    df_sorted_eff = df.sort_values('Efectividad', ascending=True)
    
    colors_eff = [get_color(x) for x in df_sorted_eff['Efectividad']]
    
    fig_eff = go.Figure()
    fig_eff.add_trace(go.Bar(
        y=df_sorted_eff['Asesor'],
        x=df_sorted_eff['Efectividad'],
        orientation='h',
        marker=dict(
            color=colors_eff,
            line=dict(color='white', width=2),
            opacity=0.85
        ),
        text=df_sorted_eff['Efectividad'].apply(lambda x: f'{x}%'),
        textposition='outside',
        textfont=dict(size=13, color='#1e293b', family='Arial', weight='bold'),
        hovertemplate='<b>%{y}</b><br><b>Efectividad:</b> <b>%{x}%</b><extra></extra>',
        name=''
    ))
    fig_eff.update_layout(
        height=580,
        margin=dict(l=50, r=50, t=20, b=20),
        showlegend=False,
        xaxis_title="",
        xaxis=dict(
            gridcolor='rgba(0,0,0,0.05)',
            showgrid=True,
            zeroline=False,
            tickfont=dict(size=11, color='#64748b'),
            range=[0, 130],
            ticksuffix='%'
        ),
        yaxis=dict(
            tickfont=dict(size=9, color='#1e293b'),
            automargin=True
        ),
        plot_bgcolor='rgba(248, 250, 252, 0.5)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=11, family='Arial', color='#1e293b'),
        hovermode='closest'
    )
    st.plotly_chart(fig_eff, use_container_width=True, config={'displayModeBar': False})

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ============= ANÁLISIS DE INSTALADAS POR SEMANA =============
st.markdown("### 📊 Análisis de Instaladas por Semana")

# Crear dos tabs: uno para mes individual y otro para comparativo
tab1, tab2 = st.tabs(["Análisis por Semana (Mes)", "Comparativo Multi-Mes"])

# TAB 1: Análisis de semanas para un mes seleccionado
with tab1:
    # Obtener meses disponibles
    meses_disp = get_meses_disponibles()
    
    if meses_disp:
        # Crear lista de opciones con formato "Mes Año"
        opciones_meses = [mes_año for mes_año, _, _, _ in meses_disp]
        
        # Obtener el mes más reciente disponible en DRIVE
        mes_reciente_tup = get_mes_mas_reciente()
        mes_reciente_display = mes_reciente_tup[0]  # Formato "Mes Año"
        
        # Buscar el índice del mes más reciente en la lista
        try:
            index_default = opciones_meses.index(mes_reciente_display)
        except ValueError:
            index_default = len(opciones_meses) - 1  # Usar el último mes disponible
        
        col_mes_sel, col_espacio = st.columns([2, 3])
        with col_mes_sel:
            mes_seleccionado_display = st.selectbox(
                "Selecciona un mes para analizar:",
                opciones_meses,
                index=index_default,  # Mes más reciente por defecto
                key="mes_analisis"
            )
        
        # Encontrar el mes_nombre del mes seleccionado
        mes_nombre_analisis = next((mes_nombre for mes_año, mes_nombre, _, _ in meses_disp if mes_año == mes_seleccionado_display), None)
    else:
        st.warning("No hay datos disponibles en los registros")
        mes_nombre_analisis = None
    
    if mes_nombre_analisis:
        # Obtener datos de instaladas por día
        df_semanas = get_instaladas_por_semana(mes_nombre_analisis)
    
    if not df_semanas.empty and len(df_semanas) > 0:
        # Crear gráfico de barras
        fig_semanas = go.Figure()
        
        fig_semanas.add_trace(go.Bar(
            x=df_semanas['DIA'],
            y=df_semanas['INSTALADAS'],
            marker=dict(
                color=df_semanas['INSTALADAS'],
                colorscale='Blues',
                line=dict(color='white', width=2),
                opacity=0.85,
                showscale=True,
                colorbar=dict(
                    title="Ventas",
                    tickfont=dict(size=10),
                    thickness=15,
                    len=0.7
                )
            ),
            text=df_semanas['INSTALADAS'],
            textposition='outside',
            textfont=dict(size=13, color='#1e293b', family='Arial', weight='bold'),
            hovertemplate='<b>%{x}</b><br><b>Ventas:</b> <b>%{y}</b><extra></extra>',
            name='Ventas'
        ))
        
        fig_semanas.update_layout(
            title=dict(
                text=f"Distribución Diaria de Ventas - {mes_seleccionado_display}",
                font=dict(size=16, color='#1e293b', family='Arial'),
                x=0.5,
                xanchor='center'
            ),
            height=550,
            margin=dict(l=50, r=50, t=80, b=150),
            xaxis_title="Día",
            yaxis_title="Cantidad de Ventas",
            xaxis=dict(
                tickfont=dict(size=9, color='#64748b'),
                tickangle=-90,
            ),
            yaxis=dict(
                gridcolor='rgba(0,0,0,0.05)',
                showgrid=True,
                zeroline=False,
                tickfont=dict(size=11, color='#64748b'),
            ),
            plot_bgcolor='rgba(248, 250, 252, 0.5)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=11, family='Arial', color='#1e293b'),
            hovermode='x unified',
            showlegend=False
        )
        
        st.plotly_chart(fig_semanas, use_container_width=True, config={'displayModeBar': False})
        
        # ============= COMPARATIVA HISTÓRICA POR DÍA: Últimos Meses =============
        st.markdown('<div style="margin-top: 40px;"></div>', unsafe_allow_html=True)
        st.markdown("#### 📊 Comparativa Histórica por Día (Últimos Meses)")
        
        df_comparativo_diario = get_comparativo_diario_multiples_meses()
        
        if not df_comparativo_diario.empty:
            # Extraer columnas de meses (sin la columna 'Día')
            meses_cols = [col for col in df_comparativo_diario.columns if col != 'Día']
            
            # Tabla comparativa detallada
            st.markdown('<div style="text-align: center;"><h4 style="margin-bottom: 20px;">📋 Tabla Comparativa Detallada</h4></div>', unsafe_allow_html=True)
            
            df_display = df_comparativo_diario.copy()
            df_display['Día'] = df_display['Día'].astype(int)
            
            # Crear tabla HTML con valores y flechas de tendencia
            html_table = '<table style="width:100%; border-collapse: collapse; margin: 0 auto;">'
            
            # Encabezado
            html_table += '<tr style="background-color: #f0f2f6; border-bottom: 2px solid #ddd;">'
            html_table += '<th style="padding: 12px; text-align: center; border-right: 1px solid #ddd;">Día</th>'
            for mes in meses_cols:
                html_table += f'<th style="padding: 12px; text-align: center; border-right: 1px solid #ddd;">{mes}</th>'
            html_table += '</tr>'
            
            # Filas de datos
            for idx, row in df_comparativo_diario.iterrows():
                dia = int(row['Día'])
                html_table += f'<tr style="border-bottom: 1px solid #eee;">'
                html_table += f'<td style="padding: 12px; text-align: center; font-weight: bold; border-right: 1px solid #eee;">{dia}</td>'
                
                for col_idx, mes in enumerate(meses_cols):
                    valor_actual = int(row[mes])
                    
                    # Determinar flecha comparando con mes anterior
                    if col_idx == 0:
                        flecha = ""
                    else:
                        mes_anterior = meses_cols[col_idx - 1]
                        valor_anterior = int(row[mes_anterior])
                        
                        if valor_actual > valor_anterior:
                            flecha = "📈"
                        elif valor_actual < valor_anterior:
                            flecha = "📉"
                        else:
                            flecha = "➡️"
                    
                    # Determinar color según tendencia
                    if flecha == "📈":
                        color = "#10b981"  # Verde
                    elif flecha == "📉":
                        color = "#ef4444"  # Rojo
                    else:
                        color = "#6b7280"  # Gris
                    
                    if flecha:
                        html_table += f'<td style="padding: 12px; text-align: center; border-right: 1px solid #eee;"><span style="color: {color};">{valor_actual} {flecha}</span></td>'
                    else:
                        html_table += f'<td style="padding: 12px; text-align: center; border-right: 1px solid #eee;">{valor_actual}</td>'
                
                html_table += '</tr>'
            
            html_table += '</table>'
            
            st.markdown(html_table, unsafe_allow_html=True)
            
            # Análisis de variación
            st.markdown("**Análisis de Variación:**")
            mes_actual = meses_cols[-1]  # Último mes (más reciente)
            mes_anterior = meses_cols[-2] if len(meses_cols) > 1 else None
            
            if mes_anterior:
                # Comparar mismo día en ambos meses
                dias_comunes = min(len(df_comparativo_diario[mes_actual]), len(df_comparativo_diario[mes_anterior]))
                
                ventas_actual_comunes = df_comparativo_diario[mes_actual][:dias_comunes].sum()
                ventas_anterior_comunes = df_comparativo_diario[mes_anterior][:dias_comunes].sum()
                
                if ventas_anterior_comunes > 0:
                    variacion_pct = ((ventas_actual_comunes - ventas_anterior_comunes) / ventas_anterior_comunes * 100)
                    variacion_texto = f"📈 +{variacion_pct:.0f}%" if variacion_pct > 0 else f"📉 {variacion_pct:.0f}%" if variacion_pct < 0 else "➡️ 0%"
                    
                    col_var1, col_var2, col_var3 = st.columns(3)
                    with col_var1:
                        st.metric(f"Ventas {mes_actual} (primeros {dias_comunes} días)", ventas_actual_comunes)
                    with col_var2:
                        st.metric(f"Ventas {mes_anterior} (primeros {dias_comunes} días)", ventas_anterior_comunes)
                    with col_var3:
                        st.metric(f"Variación", variacion_texto)
        else:
            st.info("No hay datos suficientes para realizar la comparativa histórica")
        
        # ============= COMPARATIVA POR HORARIO: FULL TIME vs PART TIME dentro del tab1 =============
        st.markdown('<div style="margin-top: 40px;"></div>', unsafe_allow_html=True)
        st.markdown("#### 📊 Comparativa por Horario: FULL TIME vs PART TIME")
        
        # Función para generar datos detallados por horario
        def generar_tabla_horario(mes_sel, meta_minima=60):
            df_lista = load_lista_metas()
            df_drive = load_drive_data()
            
            if df_lista is None or df_drive is None:
                return None, None
            
            # Preparar LISTA - limpiar espacios en Asesor y convertir Meta a numérico
            df_lista_clean = df_lista.copy()
            df_lista_clean['Asesor'] = df_lista_clean['Asesor'].astype(str).str.strip()
            df_lista_clean['Meta'] = pd.to_numeric(df_lista_clean['Meta'], errors='coerce').fillna(0)
            
            # Preparar DRIVE
            df_drive_clean = df_drive.copy()
            df_drive_clean['ASESOR'] = df_drive_clean['ASESOR'].astype(str).str.strip()
            df_drive_clean['ESTADO'] = df_drive_clean['ESTADO'].astype(str).str.strip()
            df_mes_drive = df_drive_clean[df_drive_clean['MES'] == mes_sel]
            
            # Obtener datos de LISTA
            df_mes_lista = df_lista_clean[df_lista_clean['Mes'] == mes_sel]
            
            # Clasificar asesores por horario: FULL TIME meta >= 50, PART TIME meta < 50
            full_time = df_mes_lista[df_mes_lista['Meta'] >= 50]['Asesor'].tolist()
            part_time = df_mes_lista[df_mes_lista['Meta'] < 50]['Asesor'].tolist()
            
            def procesar_horario(lista_asesoras):
                datos_tabla = []
                total_meta = 0
                total_instalado = 0
                total_pendiente = 0
                
                for idx, asesor in enumerate(lista_asesoras, 1):
                    # Meta
                    meta = df_mes_lista[df_mes_lista['Asesor'] == asesor]['Meta'].sum()
                    if meta == 0:
                        meta = 0
                    
                    # Instalado y Pendiente
                    df_asesor = df_mes_drive[df_mes_drive['ASESOR'] == asesor]
                    instalado = len(df_asesor[df_asesor['ESTADO'] == 'INSTALADO'])
                    pendiente = len(df_asesor[df_asesor['ESTADO'] == 'PENDIENTE'])
                    
                    # Alcance (Cumplimiento)
                    alcance = round((instalado / meta * 100)) if meta > 0 else 0
                    
                    datos_tabla.append({
                        'pos': idx,
                        'asesor': asesor,
                        'meta': int(meta),
                        'instalado': int(instalado),
                        'pendiente': int(pendiente),
                        'alcance': int(alcance)
                    })
                    
                    total_meta += meta
                    total_instalado += instalado
                    total_pendiente += pendiente
                
                # Ordenar datos por alcance de mayor a menor
                datos_tabla_ordenado = sorted(datos_tabla, key=lambda x: x['alcance'], reverse=True)
                
                # Actualizar posiciones después del ordenamiento
                for idx, item in enumerate(datos_tabla_ordenado, 1):
                    item['pos'] = idx
                
                return {
                    'datos': datos_tabla_ordenado,
                    'totales': {
                        'meta': int(total_meta),
                        'instalado': int(total_instalado),
                        'pendiente': int(total_pendiente),
                        'alcance': round((total_instalado / total_meta * 100)) if total_meta > 0 else 0
                    }
                }
            
            datos_full_time = procesar_horario(full_time)
            datos_part_time = procesar_horario(part_time)
            
            return datos_full_time, datos_part_time
        
        # Generar datos de ambos horarios
        datos_full_time, datos_part_time = generar_tabla_horario(mes_nombre_analisis)
        
        # Crear tablas HTML detalladas
        def crear_tabla_html_horario(datos_horario, nombre_horario, color_header, color_accent):
            if datos_horario is None or not datos_horario['datos']:
                return ""
            
            html = f'''<div style="margin: 20px 0; background: white; border-radius: 8px; overflow: auto; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
            <table style="width: 100%; border-collapse: collapse; font-family: Arial, sans-serif; min-width: 100%;">
            <thead>
                <tr style="background: {color_header}; color: white;">
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">Nº ASESOR</th>
                    <th style="padding: 12px; text-align: left; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2); min-width: 170px;">ASESOR</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">OBJETIVO</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">INSTALADO</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">PENDIENTE</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">% ALCANCE</th>
                </tr>
            </thead>
            <tbody>
            '''
            
            # Agregar filas de datos
            for item in datos_horario['datos']:
                color_fila = '#f9fafb' if item['pos'] % 2 == 0 else '#ffffff'
                alcance_color = '#10b981' if item['alcance'] >= 70 else '#f59e0b' if item['alcance'] >= 50 else '#ef4444'
                
                html += f'''<tr style="background-color: {color_fila}; border-bottom: 1px solid #e5e7eb;">
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; color: {color_accent};">#{item['pos']}</td>
                    <td style="padding: 10px 12px; text-align: left; font-weight: 500; font-size: 11px;">{item['asesor']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px;">{item['meta']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; color: #10b981;">{item['instalado']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; color: #f59e0b;">{item['pendiente']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; background-color: {alcance_color}22; color: {alcance_color}; border-radius: 4px;">{item['alcance']}%</td>
                </tr>'''
            
            # Agregar fila de totales
            totales = datos_horario['totales']
            alcance_total_color = '#10b981' if totales['alcance'] >= 70 else '#f59e0b' if totales['alcance'] >= 50 else '#ef4444'
            html += f'''<tr style="background: {color_header}; color: white; font-weight: 700;">
                    <td colspan="2" style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['meta']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['meta']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['instalado']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['pendiente']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white; background-color: {alcance_total_color}40; border-radius: 4px;">{totales['alcance']}%</td>
                </tr>
            </tbody>
            </table>
            </div>'''
            
            return html
        
        # Generar HTML para ambos horarios
        html_full_time = crear_tabla_html_horario(datos_full_time, "FULL TIME", "#3b82f6", "#3b82f6")
        html_part_time = crear_tabla_html_horario(datos_part_time, "PART TIME", "#8b5cf6", "#8b5cf6")
        
        # Mostrar tablas en dos columnas
        col_full_time, col_part_time = st.columns(2)
        
        with col_full_time:
            st.markdown('<h4 style="text-align: center; color: #3b82f6; margin-bottom: 10px;">⏰ FULL TIME </h4>', unsafe_allow_html=True)
            if html_full_time:
                st.markdown(html_full_time, unsafe_allow_html=True)
            else:
                st.info("No hay asesores FULL TIME")
        
        with col_part_time:
            st.markdown('<h4 style="text-align: center; color: #8b5cf6; margin-bottom: 10px;">⏰ PART TIME </h4>', unsafe_allow_html=True)
            if html_part_time:
                st.markdown(html_part_time, unsafe_allow_html=True)
            else:
                st.info("No hay asesores PART TIME")
    else:
        st.warning(f"No hay datos de instaladas para {mes_seleccionado_display}")

# TAB 2: Comparativo acumulativo entre meses
with tab2:
    st.markdown("#### 📈 Comparativa Acumulativa de Instaladas (Todos los Meses)")
    
    df_comparativo = get_comparativo_acumulativo_multiples_meses()
    
    if not df_comparativo.empty:
        # Crear gráfico de líneas para comparar meses
        fig_comparativo = go.Figure()
        
        # Agregar línea por cada mes
        for mes_col in df_comparativo.columns:
            fig_comparativo.add_trace(go.Scatter(
                x=df_comparativo.index,
                y=df_comparativo[mes_col],
                mode='lines+markers',
                name=mes_col,
                line=dict(width=2.5),
                marker=dict(size=6),
                hovertemplate='<b>Día %{x}</b><br><b>' + mes_col + ':</b> %{y} acumuladas<extra></extra>'
            ))
        
        fig_comparativo.update_layout(
            title=dict(
                text="Comparativa Acumulativa de Instaladas por Día (Todos los Meses)",
                font=dict(size=16, color='#1e293b', family='Arial'),
                x=0.5,
                xanchor='center'
            ),
            height=550,
            margin=dict(l=60, r=60, t=80, b=80),
            xaxis_title="Día del Mes",
            yaxis_title="Total Acumulado de Instaladas",
            xaxis=dict(
                tickfont=dict(size=11, color='#64748b'),
                tickmode='linear',
                tick0=1,
                dtick=2
            ),
            yaxis=dict(
                gridcolor='rgba(0,0,0,0.05)',
                showgrid=True,
                zeroline=False,
                tickfont=dict(size=11, color='#64748b'),
            ),
            plot_bgcolor='rgba(248, 250, 252, 0.5)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=11, family='Arial', color='#1e293b'),
            hovermode='x unified',
            legend=dict(
                x=1.02,
                y=1,
                xanchor='left',
                yanchor='top',
                bgcolor='rgba(255, 255, 255, 0.8)',
                bordercolor='#e2e8f0',
                borderwidth=1
            )
        )
        
        st.plotly_chart(fig_comparativo, use_container_width=True, config={'displayModeBar': False})
    else:
        st.warning("No hay datos suficientes para el comparativo de meses")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Tabla de detalle de asesores
st.markdown("### 👥 Detalle Completo de Asesores")

# Filtro para ordenamiento
criterio_orden = st.selectbox(
    "Ordenar por:",
    ["Cumplimiento (Mayor a Menor)", "Conversión (Mayor a Menor)"],
    key="criterio_orden"
)

df_detail = df[['Asesor', 'Meta', 'Instaladas', 'Canceladas', 'Cumplimiento', 'Efectividad']].copy()
df_detail['Cumpl%'] = df_detail['Cumplimiento'].astype(str) + '%'
df_detail['Efect%'] = df_detail['Efectividad'].astype(str) + '%'

# Agregar columna de Pendientes solo para el mes actual
# Obtener mes actual
meses_nombres = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
                 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
mes_actual = meses_nombres[datetime.now().month]

# Solo agregar columna de Pendientes si el mes seleccionado es el mes actual
if mes == mes_actual:
    pendientes_list = []
    for asesor in df_detail['Asesor']:
        pendientes = get_pendientes_asesor_mes(asesor, mes)
        pendientes_list.append(pendientes)
    df_detail['Pendientes'] = pendientes_list

# Agregar columnas de Leads, Con Cobertura y Ventas
leads_list = []
con_cobertura_list = []
ventas_list = []
for asesor in df_detail['Asesor']:
    leads = get_leads_asesor_mes(asesor, mes)
    con_cobertura = get_con_cobertura_asesor_mes(asesor, mes)
    ventas = get_ventas_asesor_mes(asesor, mes)
    leads_list.append(leads)
    con_cobertura_list.append(con_cobertura)
    ventas_list.append(ventas)
df_detail['Leads'] = leads_list
df_detail['Con Cobertura'] = con_cobertura_list
df_detail['Ventas'] = ventas_list

# Separar en Full Time (meta >= 55) y Part Time (meta < 55)
# Excepción: ISABEL, LAURA y KARINA son FULL TIME aunque tengan meta < 55
asesoras_fulltime_especial = ['ZIM_ISABELPF_VTP', 'ZIM_LAURAVS_VTP', 'ZIM_KARINASE_VTP']
condicion_fulltime = (df_detail['Meta'] >= 55) | (df_detail['Asesor'].isin(asesoras_fulltime_especial))
df_fulltime = df_detail[condicion_fulltime].copy()
df_parttime = df_detail[~condicion_fulltime].copy()

# Ordenar SIEMPRE por Conversión (Efectividad) de mayor a menor
df_fulltime = df_fulltime.sort_values('Efectividad', ascending=False).reset_index(drop=True)
df_parttime = df_parttime.sort_values('Efectividad', ascending=False).reset_index(drop=True)

# Función para generar tabla HTML
def generar_tabla_detalle(df_tabla, tipo_empleado):
    # Obtener mes actual
    meses_nombres = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
                     7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
    mes_actual = meses_nombres[datetime.now().month]
    
    # Verificar si mostrar columna de Pendientes (solo si es el mes actual)
    mostrar_pendientes = mes == mes_actual
    
    if mostrar_pendientes:
        html_tabla = '<div class="meta-tabla"><table><thead><tr><th style="width: 3%;">Pos</th><th style="width: 15%;">Asesor</th><th style="width: 5%;">Meta</th><th style="width: 5%;">Leads</th><th style="width: 6%;">Cob</th><th style="width: 6%;">Ventas</th><th style="width: 6%;">Inst</th><th style="width: 6%;">Canc</th><th style="width: 6%;">Pend</th><th style="width: 7%;">Cumpl%</th><th style="width: 8%;">Conv%</th><th style="width: 11%;">Estado</th></tr></thead><tbody>'
    else:
        html_tabla = '<div class="meta-tabla"><table><thead><tr><th style="width: 4%;">Pos</th><th style="width: 18%;">Asesor</th><th style="width: 6%;">Meta</th><th style="width: 6%;">Leads</th><th style="width: 6%;">Cob</th><th style="width: 6%;">Ventas</th><th style="width: 7%;">Inst</th><th style="width: 7%;">Canc</th><th style="width: 8%;">Cumpl%</th><th style="width: 10%;">Conv%</th><th style="width: 10%;">Estado</th></tr></thead><tbody>'

    for idx, (_, row) in enumerate(df_tabla.iterrows(), 1):
        asesor = row['Asesor']
        leads = int(row.get('Leads', 0))
        con_cobertura = int(row.get('Con Cobertura', 0))
        ventas = int(row.get('Ventas', 0))
        meta = int(row['Meta'])
        instaladas = int(row['Instaladas'])
        canceladas = int(row['Canceladas'])
        pendientes = int(row.get('Pendientes', 0)) if mostrar_pendientes else 0
        efect = int(row['Efectividad'])
        
        # Calcular CUMPL% como Instaladas / Meta * 100
        cumpl_semana = round((instaladas / meta * 100)) if meta > 0 else 0
        
        # Determinar color basado en el cumplimiento respecto a la meta
        if cumpl_semana >= 100:
            cumpl_color = '#10b981'  # Verde - Meta alcanzada
            cumpl_bg_color = 'rgba(16, 185, 129, 0.2)'
        elif cumpl_semana >= 75:
            cumpl_color = '#059669'  # Verde más oscuro - Buen progreso
            cumpl_bg_color = 'rgba(5, 150, 105, 0.2)'
        elif cumpl_semana >= 50:
            cumpl_color = '#f59e0b'  # Naranja - En ruta pero no garantizado
            cumpl_bg_color = 'rgba(245, 158, 11, 0.2)'
        else:
            cumpl_color = '#ef4444'  # Rojo - Crítico
            cumpl_bg_color = 'rgba(239, 68, 68, 0.2)'
        
        # Determinar estado general basado en CONVERSIÓN (Efectividad)
        if efect > 70:
            estado = '<span class="status-excellent">✓ Excelente</span>'
            fila_bg = 'background-color: #f0fdf4;'
        elif efect >= 50:
            estado = '<span class="status-good">~ Regular</span>'
            fila_bg = 'background-color: #fffbeb;'
        else:
            estado = '<span class="status-poor">✗ Bajo</span>'
            fila_bg = 'background-color: #fef2f2;'
        
        if mostrar_pendientes:
            html_tabla += f'''<tr style="{fila_bg}">
                <td style="font-weight: 700; text-align: center; color: #0066cc;">#{idx}</td>
                <td style="font-weight: 600;">{asesor}</td>
                <td style="text-align: center; font-weight: 600;">{meta}</td>
                <td style="text-align: center; font-weight: 600; color: #0066cc;">{leads}</td>
                <td style="text-align: center; font-weight: 600; color: #8b5cf6;">{con_cobertura}</td>
                <td style="text-align: center; font-weight: 600; color: #059669;">{ventas}</td>
                <td style="text-align: center; font-weight: 600; color: #10b981;">{instaladas}</td>
                <td style="text-align: center; font-weight: 600; color: #ef4444;">{canceladas}</td>
                <td style="text-align: center; font-weight: 600; color: #f59e0b;">{pendientes}</td>
                <td style="text-align: center; background-color: {cumpl_bg_color}; border-radius: 4px; color: {cumpl_color}; font-weight: 700;"><div class="meta-valor" style="background: none; color: {cumpl_color};">{cumpl_semana}%</div></td>
                <td style="text-align: center;"><div class="meta-valor" style="background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);">{efect}%</div></td>
                <td style="text-align: center;">{estado}</td>
            </tr>'''
        else:
            html_tabla += f'''<tr style="{fila_bg}">
                <td style="font-weight: 700; text-align: center; color: #0066cc;">#{idx}</td>
                <td style="font-weight: 600;">{asesor}</td>
                <td style="text-align: center; font-weight: 600;">{meta}</td>
                <td style="text-align: center; font-weight: 600; color: #0066cc;">{leads}</td>
                <td style="text-align: center; font-weight: 600; color: #8b5cf6;">{con_cobertura}</td>
                <td style="text-align: center; font-weight: 600; color: #059669;">{ventas}</td>
                <td style="text-align: center; font-weight: 600; color: #10b981;">{instaladas}</td>
                <td style="text-align: center; font-weight: 600; color: #ef4444;">{canceladas}</td>
                <td style="text-align: center; background-color: {cumpl_bg_color}; border-radius: 4px; color: {cumpl_color}; font-weight: 700;"><div class="meta-valor" style="background: none; color: {cumpl_color};">{cumpl_semana}%</div></td>
                <td style="text-align: center;"><div class="meta-valor" style="background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);">{efect}%</div></td>
                <td style="text-align: center;">{estado}</td>
            </tr>'''

    html_tabla += '</tbody></table></div>'
    return html_tabla

# Mostrar tabla Full Time
if not df_fulltime.empty:
    st.markdown("#### 💼 Asesores Full Time (8 horas - Meta ≥ 55)")
    html_fulltime = generar_tabla_detalle(df_fulltime, "Full Time")
    st.markdown(html_fulltime, unsafe_allow_html=True)
    st.markdown('<div style="margin: 15px 0;"></div>', unsafe_allow_html=True)

# Mostrar tabla Part Time
if not df_parttime.empty:
    st.markdown("#### 👨‍💼 Asesores Part Time (4 horas - Meta < 55)")
    html_parttime = generar_tabla_detalle(df_parttime, "Part Time")
    st.markdown(html_parttime, unsafe_allow_html=True)

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Mostrar tabla combinada de todos los asesores por CODIGO DE CARGA
st.markdown("#### 👥 Detalle Completo de Todos los Asesores")

# Cargar datos agrupados por CODIGO DE CARGA para el mes seleccionado
df_codigos_carga = load_data_codigo_carga(mes)

if not df_codigos_carga.empty:
    # Función para calcular color basado en porcentaje
    def get_color_for_percentage(percentage):
        """
        Retorna color basado en porcentaje.
        >= 60%: Verde
        < 60%: Escala de rojo (más bajo = más intenso)
        """
        if percentage >= 60:
            return '#10b981'  # Verde
        else:
            # Escala de rojo desde claro (59%) a intenso (0%)
            # Interpolamos de #FFB3B3 (rojo claro) a #FF0000 (rojo intenso)
            normalized = (60 - percentage) / 60  # 0 cuando 60%, 1 cuando 0%
            g_value = int(179 * (1 - normalized))  # De 179 a 0
            b_value = int(179 * (1 - normalized))  # De 179 a 0
            return f'#FF{g_value:02X}{b_value:02X}'
    
    # Crear tabla HTML con el formato deseado
    def generar_tabla_codigos_carga(df_datos):
        """Genera tabla HTML para datos agrupados por CODIGO DE CARGA"""
        html = '''<div style="margin: 20px 0; background: white; border-radius: 8px; overflow: auto; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
        <table style="width: 100%; border-collapse: collapse; font-family: Arial, sans-serif;">
        <thead>
            <tr style="background: linear-gradient(135deg, #0066cc 0%, #00d4ff 100%); color: white;">
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">POS</th>
                <th style="padding: 14px; text-align: left; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2); min-width: 180px;">CODIGO CARGA</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">TOTAL DE LEADS</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);"># CON COBERTURA</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">TOTAL DE VENTAS</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">%CONV. VENTAS FINAL</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px;">%CONV. VENTAS</th>
            </tr>
        </thead>
        <tbody>
        '''
        
        for idx, row in df_datos.iterrows():
            color_fila = '#f9fafb' if idx % 2 == 0 else '#ffffff'
            pos = int(row['POS'])
            codigo = row['CODIGO_CARGA']
            leads = int(row['LEADS'])
            con_cobertura = int(row['CON_COBERTURA'])
            ventas = int(row['VENTAS'])
            conv_ventas = float(row['CONV_VENTAS'])
            conv_ventas_cob = float(row['CONV_VENTAS_COB'])
            
            # Determinar color para ventas
            if ventas > 0:
                color_ventas = '#10b981'  # Verde
            else:
                color_ventas = '#64748b'  # Gris
            
            # Determinar color para conversión respecto a leads
            color_conv = get_color_for_percentage(conv_ventas)
            
            # Determinar color para conversión respecto a con cobertura
            color_conv_cob = get_color_for_percentage(conv_ventas_cob)
            
            html += f'''<tr style="background-color: {color_fila}; border-bottom: 1px solid #e5e7eb;">
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; color: #0066cc;">#{pos}</td>
                <td style="padding: 12px; text-align: left; font-weight: 500; font-size: 12px;">{codigo}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px;">{leads}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px;">{con_cobertura}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; color: {color_ventas};">{ventas}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; background-color: {color_conv}22; color: {color_conv}; border-radius: 4px;">{conv_ventas:.2f}%</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; background-color: {color_conv_cob}22; color: {color_conv_cob}; border-radius: 4px;">{conv_ventas_cob:.2f}%</td>
            </tr>'''
        
        # Calcular y agregar fila de TOTALES
        total_leads = df_datos['LEADS'].sum()
        total_con_cobertura = df_datos['CON_COBERTURA'].sum()
        total_ventas = df_datos['VENTAS'].sum()
        total_conv_ventas_cob = round((total_ventas / total_con_cobertura * 100), 2) if total_con_cobertura > 0 else 0.0
        total_conv_ventas = round((total_ventas / total_leads * 100), 2) if total_leads > 0 else 0.0
        
        # Determinar color para conversión total respecto a con cobertura y a leads
        color_conv_total_cob = get_color_for_percentage(total_conv_ventas_cob)
        color_conv_total = get_color_for_percentage(total_conv_ventas)
        
        html += f'''<tr style="background: linear-gradient(135deg, #0066cc 0%, #00d4ff 100%); color: white; font-weight: 700;">
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;"></td>
            <td style="padding: 12px; text-align: left; font-weight: 700; font-size: 12px;">TOTAL</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">{total_leads}</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">{total_con_cobertura}</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">{total_ventas}</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; background-color: {color_conv_total}40; color: white; border-radius: 4px;">{total_conv_ventas:.2f}%</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; background-color: {color_conv_total_cob}40; color: white; border-radius: 4px;">{total_conv_ventas_cob:.2f}%</td>
        </tr>'''
        
        html += '''</tbody>
        </table>
        </div>'''
        
        return html
    
    # Generar y mostrar tabla
    html_tabla = generar_tabla_codigos_carga(df_codigos_carga)
    st.markdown(html_tabla, unsafe_allow_html=True)
    
    # ============= FILTRO Y ANÁLISIS DE ASESORES EN RIESGO =============
    st.markdown("#### 🎯 Filtro de Asesores")
    
    # Crear filtro multiselect para excluir asesores
    col_filtro1, col_filtro2 = st.columns([3, 1])
    
    with col_filtro1:
        asesores_todos = sorted(df_codigos_carga['CODIGO_CARGA'].tolist())
        asesores_excluir = st.multiselect(
            "Excluir asesores (vacaciones, descanso, dados de baja, etc.)",
            options=asesores_todos,
            default=[],
            key=f"excluir_asesores_{mes}"
        )
    
    # Filtrar dataframe excluyendo asesores seleccionados
    df_codigos_filtrado = df_codigos_carga[~df_codigos_carga['CODIGO_CARGA'].isin(asesores_excluir)].copy()
    
    # ============= CUADRO DE ASESORES EN RIESGO =============
    st.markdown("#### ⚠️ Asesores en Riesgo (Conversión < 60%)")
    
    # Filtrar asesores con conversión < 60%
    df_en_riesgo = df_codigos_filtrado[df_codigos_filtrado['CONV_VENTAS_COB'] < 60].copy()
    
    if not df_en_riesgo.empty:
        # Ordenar por conversión (de menor a mayor)
        df_en_riesgo = df_en_riesgo.sort_values('CONV_VENTAS_COB').reset_index(drop=True)
        
        def generar_tabla_riesgo(df_datos):
            """Genera tabla HTML para asesores en riesgo"""
            html = '''<div style="margin: 20px 0; background: white; border-radius: 8px; overflow: auto; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
            <table style="width: 100%; border-collapse: collapse; font-family: Arial, sans-serif;">
            <thead>
                <tr style="background: linear-gradient(135deg, #ff6b6b 0%, #ff8787 100%); color: white;">
                    <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">POSICIÓN</th>
                    <th style="padding: 14px; text-align: left; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2); min-width: 180px;">ASESOR</th>
                    <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">LEADS</th>
                    <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">CON COBERTURA</th>
                    <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">VENTAS</th>
                    <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px;">CONVERSIÓN %</th>
                </tr>
            </thead>
            <tbody>
            '''
            
            for idx, row in df_datos.iterrows():
                color_fila = '#fff5f5' if idx % 2 == 0 else '#ffffff'
                codigo = row['CODIGO_CARGA']
                leads = int(row['LEADS'])
                con_cobertura = int(row['CON_COBERTURA'])
                ventas = int(row['VENTAS'])
                conv = int(row['CONV_VENTAS_COB'])
                
                # Usamos la función get_color_for_percentage para el color
                color_conv = get_color_for_percentage(conv)
                
                html += f'''<tr style="background-color: {color_fila}; border-bottom: 1px solid #ffe0e0;">
                    <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; color: #ff6b6b;">#{idx + 1}</td>
                    <td style="padding: 12px; text-align: left; font-weight: 500; font-size: 12px;">{codigo}</td>
                    <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px;">{leads}</td>
                    <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px;">{con_cobertura}</td>
                    <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px;">{ventas}</td>
                    <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; background-color: {color_conv}33; color: {color_conv}; border-radius: 4px; font-weight: 700;">{conv}%</td>
                </tr>'''
            
            html += '''</tbody>
            </table>
            </div>'''
            return html
        
        html_riesgo = generar_tabla_riesgo(df_en_riesgo)
        st.markdown(html_riesgo, unsafe_allow_html=True)
        
        # Mostrar estadísticas de asesores en riesgo
        st.markdown("**📊 Estadísticas de Asesores en Riesgo:**")
        col_est1, col_est2, col_est3, col_est4 = st.columns(4)
        
        with col_est1:
            st.metric("Total en Riesgo", len(df_en_riesgo))
        with col_est2:
            st.metric("Promedio Conversión", f"{df_en_riesgo['CONV_VENTAS_COB'].mean():.0f}%")
        with col_est3:
            st.metric("Más Bajo", f"{df_en_riesgo['CONV_VENTAS_COB'].min()}%")
        with col_est4:
            st.metric("Total Leads", int(df_en_riesgo['LEADS'].sum()))
    else:
        st.success("✅ ¡Excelente! No hay asesores en riesgo. Todos superan el 60% de conversión.")
    
    # Mostrar estadísticas generales
    st.markdown("#### 📊 Resumen General por Mes")
    col1, col2, col3, col4, col5 = st.columns(5)
    
    total_codigos = len(df_codigos_carga)
    total_leads = df_codigos_carga['LEADS'].sum()
    total_ventas = df_codigos_carga['VENTAS'].sum()
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Códigos de Carga", total_codigos)
    with col2:
        st.metric("Total Leads", total_leads)
    with col3:
        st.metric("Total Ventas", total_ventas)
else:
    st.info("No hay datos disponibles para el mes seleccionado")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ============= NUEVA SECCIÓN: ANÁLISIS DE CASOS POR NIVEL =============
st.markdown("### 📋 Análisis de Casos por Nivel (MANTRA)")

# Obtener datos detallados de MANTRA
df_mantra_mes = get_datos_mantra_mes(mes)

if not df_mantra_mes.empty:
    # Crear filtros múltiples en 4 columnas
    col1, col2, col3, col4 = st.columns(4, gap="small")
    
    with col1:
        agentes_unique = sorted(df_mantra_mes['Agente'].unique())
        agente_filtro = st.selectbox(
            "Agente",
            ["Todos"] + list(agentes_unique),
            key="agente_filtro_casos"
        )
    
    # Filtrar por agente para obtener valores únicos de niveles
    if agente_filtro == "Todos":
        df_temp = df_mantra_mes.copy()
    else:
        df_temp = df_mantra_mes[df_mantra_mes['Agente'] == agente_filtro]
    
    with col2:
        if not df_temp.empty and 'NIVEL 1' in df_temp.columns:
            nivel1_values = df_temp['NIVEL 1'].dropna().unique()
            nivel1_unique = sorted([str(x).strip() for x in nivel1_values if str(x).strip() and str(x).lower() != 'nan'])
        else:
            nivel1_unique = []
        
        nivel1_filtro = st.selectbox(
            "Nivel 1",
            ["Todos"] + list(nivel1_unique),
            key="nivel1_filtro_casos"
        )
    
    # Filtrar por nivel 1
    if nivel1_filtro == "Todos":
        df_temp2 = df_temp.copy()
    else:
        if not df_temp.empty:
            df_temp2 = df_temp[df_temp['NIVEL 1'] == nivel1_filtro]
        else:
            df_temp2 = pd.DataFrame()
    
    with col3:
        if not df_temp2.empty and 'NIVEL 2' in df_temp2.columns:
            nivel2_values = df_temp2['NIVEL 2'].dropna().unique()
            nivel2_unique = sorted([str(x).strip() for x in nivel2_values if str(x).strip() and str(x).lower() != 'nan'])
        else:
            nivel2_unique = []
        
        nivel2_filtro = st.selectbox(
            "Nivel 2",
            ["Todos"] + list(nivel2_unique),
            key="nivel2_filtro_casos"
        )
    
    # Filtrar por nivel 2
    if nivel2_filtro == "Todos":
        df_temp3 = df_temp2.copy()
    else:
        if not df_temp2.empty:
            df_temp3 = df_temp2[df_temp2['NIVEL 2'] == nivel2_filtro]
        else:
            df_temp3 = pd.DataFrame()
    
    with col4:
        if not df_temp3.empty and 'NIVEL 3' in df_temp3.columns:
            nivel3_values = df_temp3['NIVEL 3'].dropna().unique()
            nivel3_unique = sorted([str(x).strip() for x in nivel3_values if str(x).strip() and str(x).lower() != 'nan'])
        else:
            nivel3_unique = []
        
        nivel3_filtro = st.multiselect(
            "Nivel 3",
            list(nivel3_unique),
            default=list(nivel3_unique),
            key="nivel3_filtro_casos"
        )
    
    # Segunda fila de filtros - Fechas
    col_fecha1, col_fecha2 = st.columns(2, gap="small")
    
    # Preparar datos de fecha
    fecha_filtrada = df_mantra_mes.copy()
    if 'FECHA' in fecha_filtrada.columns or 'Fecha' in fecha_filtrada.columns:
        # Detectar columna de fecha
        col_fecha = 'FECHA' if 'FECHA' in fecha_filtrada.columns else 'Fecha'
        fecha_filtrada[col_fecha] = pd.to_datetime(fecha_filtrada[col_fecha], errors='coerce')
        
        # Obtener rango de fechas disponibles
        fecha_min = fecha_filtrada[col_fecha].min()
        fecha_max = fecha_filtrada[col_fecha].max()
        
        with col_fecha1:
            fecha_inicio = st.date_input(
                "Fecha Inicio",
                value=fecha_min,
                min_value=fecha_min,
                max_value=fecha_max,
                key="fecha_inicio_casos"
            )
        
        with col_fecha2:
            fecha_fin = st.date_input(
                "Fecha Fin",
                value=fecha_max,
                min_value=fecha_min,
                max_value=fecha_max,
                key="fecha_fin_casos"
            )
    else:
        fecha_inicio = None
        fecha_fin = None
    
    # Aplicar todos los filtros
    df_filtrado = df_mantra_mes.copy()
    
    if agente_filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado['Agente'] == agente_filtro]
    
    if nivel1_filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado['NIVEL 1'] == nivel1_filtro]
    
    if nivel2_filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado['NIVEL 2'] == nivel2_filtro]
    
    if nivel3_filtro:
        df_filtrado = df_filtrado[df_filtrado['NIVEL 3'].isin(nivel3_filtro)]
    
    # Aplicar filtro de fechas si existen
    if fecha_inicio is not None and fecha_fin is not None:
        col_fecha = 'FECHA' if 'FECHA' in df_filtrado.columns else 'Fecha'
        if col_fecha in df_filtrado.columns:
            df_filtrado[col_fecha] = pd.to_datetime(df_filtrado[col_fecha], errors='coerce')
            df_filtrado = df_filtrado[
                (df_filtrado[col_fecha] >= pd.Timestamp(fecha_inicio)) &
                (df_filtrado[col_fecha] <= pd.Timestamp(fecha_fin))
            ]
    
    # Mostrar total de casos filtrados
    total_casos_filtrados = len(df_filtrado)
    st.markdown(f"### Total de Casos: **{total_casos_filtrados}**")
    
    # Mostrar vista previa (primeros 10 registros)
    if total_casos_filtrados > 0:
        col_preview, col_download = st.columns([4, 1])
        
        with col_preview:
            st.markdown("#### Vista Previa (Primeros 10 registros)")
            
            # Seleccionar columnas para mostrar
            cols_mostrar = ['Agente', 'NIVEL 1', 'NIVEL 2', 'NIVEL 3']
            if 'Telefono' in df_filtrado.columns:
                cols_mostrar.insert(1, 'Telefono')
            if 'Numero Caso' in df_filtrado.columns:
                cols_mostrar.insert(0, 'Numero Caso')
            
            df_preview = df_filtrado[cols_mostrar].head(10)
            st.dataframe(df_preview, use_container_width=True, hide_index=True)
            
            if total_casos_filtrados > 10:
                st.caption(f"Mostrando 10 de {total_casos_filtrados} registros")
        
        with col_download:
            st.markdown("#### Descargar")
            
            # Crear archivo Excel con todos los datos filtrados
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_filtrado.to_excel(writer, sheet_name='Casos Filtrados', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Casos Filtrados']
                
                # Estilos
                header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                
                # Aplicar estilos al encabezado
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Aplicar estilos a datos
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Ajustar anchos de columna
                for col_num, col in enumerate(worksheet.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(col_num)
                    worksheet.column_dimensions[col_letter].width = 20
            
            buffer.seek(0)
            
            # Preparar nombre del archivo con agente y fecha
            nombre_asesor = agente_filtro if agente_filtro != "Todos" else "Todos"
            fecha_hoy = datetime.now().strftime('%d_%m_%Y')
            nombre_archivo = f"Casos_Filtrados_{nombre_asesor}_{mes}_{fecha_hoy}.xlsx"
            
            st.download_button(
                label="📥 Descargar Excel",
                data=buffer,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.warning("No hay casos que coincidan con los filtros seleccionados")

else:
    st.warning(f"No hay datos de casos disponibles para {mes}")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ============= NUEVA SECCIÓN: ANÁLISIS DETALLADO DEL DRIVE =============
st.markdown("### 📊 Análisis Detallado del DRIVE por Asesor")
st.markdown("*Historial de ventas, comportamiento y recomendaciones personalizadas*")

# Obtener lista de asesores del DRIVE para el mes
df_drive_mes_actual = load_drive_data()
if df_drive_mes_actual is not None and not df_drive_mes_actual.empty:
    df_drive_mes_actual = df_drive_mes_actual[df_drive_mes_actual['MES'] == mes].copy()
    df_drive_mes_actual['ASESOR'] = df_drive_mes_actual['ASESOR'].astype(str).str.strip()
    asesores_drive = sorted(df_drive_mes_actual['ASESOR'].unique())
    col1, col2 = st.columns([3, 1])
    
    with col1:
        asesor_seleccionado = st.selectbox(
            "Selecciona un asesor para análisis detallado:",
            asesores_drive,
            key="asesor_drive_analysis"
        )
    
    # Obtener datos del asesor
    kpis = get_drive_asesor_kpis(asesor_seleccionado, mes)
    
    if kpis:
        # FILA 1: KPIs PRINCIPALES
        st.markdown("#### 📈 Métricas Clave del Mes")
        
        metric_cols = st.columns(5)
        
        with metric_cols[0]:
            st.metric(
                "Total Ventas",
                int(kpis['total_ventas']),
                help="Todas las transacciones registradas"
            )
        
        with metric_cols[1]:
            st.metric(
                "Instaladas",
                int(kpis['instaladas']),
                delta=f"{kpis['tasa_conversion']:.0f}%",
                help="Ventas efectivas"
            )
        
        with metric_cols[2]:
            color_cancelacion = "🔴" if kpis['tasa_cancelacion'] > 30 else "🟡" if kpis['tasa_cancelacion'] > 15 else "🟢"
            st.metric(
                "Canceladas",
                f"{color_cancelacion} {int(kpis['canceladas'])}",
                delta=f"-{kpis['tasa_cancelacion']:.0f}%",
                help="Ventas que se cancelaron"
            )
        
        with metric_cols[3]:
            st.metric(
                "Pendientes",
                int(kpis['pendientes']),
                help="Ventas en seguimiento"
            )
        
        with metric_cols[4]:
            st.metric(
                "Velocidad",
                f"{kpis['velocidad_venta']:.0f} /día",
                help="Ventas por día promedio"
            )
        
        # FILA 2C: GRÁFICA DE CRECIMIENTO DE VENTAS
        st.markdown("#### 📈 Crecimiento de Ventas por Fecha")
        
        # Tabs para seleccionar vista por Día, Semana o Análisis de Metas
        tab_dia, tab_semana, tab_metas = st.tabs(["📅 Por Día", "📊 Por Semana", "🎯 Análisis de Metas"])
        
        # TAB 1: Visualización por Día
        with tab_dia:
            df_crecimiento = get_crecimiento_ventas(asesor_seleccionado, mes)
            
            if not df_crecimiento.empty:
                # Gráfico de línea de crecimiento acumulado CON PROMEDIO
                fig_crecimiento = go.Figure()
                
                # Línea de total acumulado
                fig_crecimiento.add_trace(go.Scatter(
                    x=df_crecimiento['Fecha'],
                    y=df_crecimiento['Total Acumulado'],
                    mode='lines+markers',
                    name='Total Acumulado',
                    line=dict(color='#1976d2', width=3),
                    marker=dict(size=8)
                ))
                
                # Línea de instaladas acumuladas
                fig_crecimiento.add_trace(go.Scatter(
                    x=df_crecimiento['Fecha'],
                    y=df_crecimiento['Instaladas Acumuladas'],
                    mode='lines+markers',
                    name='Instaladas Acumuladas',
                    line=dict(color='#4caf50', width=2, dash='dash'),
                    marker=dict(size=6)
                ))
                
                fig_crecimiento.update_layout(
                    title=f"Crecimiento Acumulado de Ventas por Día - {mes}",
                    xaxis_title="Fecha",
                    yaxis_title="Cantidad Acumulada",
                    height=400,
                    hovermode='x unified',
                    plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray'),
                    yaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray')
                )
                
                st.plotly_chart(fig_crecimiento, use_container_width=True)
                
                # Gráfico de barras: Desempeño diario vs promedio
                promedio_diario = df_crecimiento['TOTAL'].mean()
                colores_barras = ['#4caf50' if x >= promedio_diario else '#ff6b6b' for x in df_crecimiento['TOTAL']]
                
                fig_desempeño = go.Figure(data=[
                    go.Bar(
                        x=df_crecimiento['Fecha'],
                        y=df_crecimiento['TOTAL'],
                        marker=dict(color=colores_barras),
                        name='Ventas del Día',
                        text=df_crecimiento['TOTAL'],
                        textposition='auto'
                    )
                ])
                
                # Agregar línea de promedio
                fig_desempeño.add_hline(
                    y=promedio_diario,
                    line_dash="dash",
                    line_color="orange",
                    annotation_text=f"Promedio: {promedio_diario:.0f}",
                    annotation_position="right"
                )
                
                fig_desempeño.update_layout(
                    title=f"Desempeño Diario vs Promedio - {mes}",
                    xaxis_title="Fecha",
                    yaxis_title="Ventas del Día",
                    height=400,
                    hovermode='x unified',
                    plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray'),
                    yaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray'),
                    showlegend=False
                )
                
                st.plotly_chart(fig_desempeño, use_container_width=True)
                
                # Mostrar indicadores de crecimiento
                if len(df_crecimiento) > 1:
                    inicio = df_crecimiento['Total Acumulado'].iloc[0]
                    fin = df_crecimiento['Total Acumulado'].iloc[-1]
                    crecimiento_total = fin - inicio
                    dias_trabajados = len(df_crecimiento)
                    crecimiento_promedio_diario = crecimiento_total / dias_trabajados if dias_trabajados > 0 else 0
                    
                    # Calcular velocidad reciente (últimos 3 días)
                    if len(df_crecimiento) >= 3:
                        velocidad_reciente = df_crecimiento['TOTAL'].iloc[-3:].mean()
                    else:
                        velocidad_reciente = df_crecimiento['TOTAL'].mean()
                    
                    # Contar días arriba y bajo promedio
                    dias_arriba = len(df_crecimiento[df_crecimiento['TOTAL'] >= promedio_diario])
                    dias_bajo = len(df_crecimiento[df_crecimiento['TOTAL'] < promedio_diario])
                    
                    col_crec1, col_crec2, col_crec3 = st.columns(3)
                    with col_crec1:
                        st.metric("Crecimiento Total", f"+{crecimiento_total:.0f} ventas")
                    with col_crec2:
                        st.metric("Promedio Diario", f"{promedio_diario:.0f} ventas/día")
                    with col_crec3:
                        st.metric("Velocidad Reciente", f"{velocidad_reciente:.0f} ventas/día")
                    
                    # Segunda fila de métricas
                    col_crec4, col_crec5 = st.columns(2)
                    with col_crec4:
                        st.metric("🟢 Días Arriba del Promedio", f"{dias_arriba}/{dias_trabajados}")
                    with col_crec5:
                        st.metric("🔴 Días Bajo el Promedio", f"{dias_bajo}/{dias_trabajados}")
            else:
                st.info("No hay datos de crecimiento por día para este mes")
        
        # TAB 2: Visualización por Semana
        with tab_semana:
            df_crecimiento_semanal = get_crecimiento_ventas_semanal(asesor_seleccionado, mes)
            
            if not df_crecimiento_semanal.empty:
                # Gráfico de barras: Conteo de PAGO por semana
                promedio_semanal = df_crecimiento_semanal['TOTAL'].mean()
                colores_barras_sem = ['#4caf50' if x >= promedio_semanal else '#ff6b6b' for x in df_crecimiento_semanal['TOTAL']]
                
                fig_desempeño_sem = go.Figure(data=[
                    go.Bar(
                        x=df_crecimiento_semanal['Semana'],
                        y=df_crecimiento_semanal['TOTAL'],
                        marker=dict(color=colores_barras_sem),
                        name='PAGO por Semana',
                        text=df_crecimiento_semanal['TOTAL'],
                        textposition='auto'
                    )
                ])
                
                # Agregar línea de promedio
                fig_desempeño_sem.add_hline(
                    y=promedio_semanal,
                    line_dash="dash",
                    line_color="orange",
                    annotation_text=f"Promedio: {promedio_semanal:.0f}",
                    annotation_position="right"
                )
                
                fig_desempeño_sem.update_layout(
                    title=f"Conteo de PAGO por Semana - {mes} ({asesor_seleccionado})",
                    xaxis_title="Semana",
                    yaxis_title="Cantidad de PAGO",
                    height=400,
                    hovermode='x unified',
                    plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray'),
                    yaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray'),
                    showlegend=False
                )
                
                st.plotly_chart(fig_desempeño_sem, use_container_width=True)
                
                # Mostrar indicadores semanal
                if len(df_crecimiento_semanal) > 0:
                    total_pagos = df_crecimiento_semanal['TOTAL'].sum()
                    semanas_con_datos = len(df_crecimiento_semanal)
                    
                    # Velocidad más reciente (última semana)
                    velocidad_semana_actual = df_crecimiento_semanal['TOTAL'].iloc[-1]
                    
                    # Contar semanas arriba y bajo promedio
                    semanas_arriba = len(df_crecimiento_semanal[df_crecimiento_semanal['TOTAL'] >= promedio_semanal])
                    semanas_bajo = len(df_crecimiento_semanal[df_crecimiento_semanal['TOTAL'] < promedio_semanal])
                    
                    col_sem1, col_sem2, col_sem3 = st.columns(3)
                    with col_sem1:
                        st.metric("Total de PAGO", f"{total_pagos:.0f}")
                    with col_sem2:
                        st.metric("Promedio por Semana", f"{promedio_semanal:.0f} PAGO/semana")
                    with col_sem3:
                        st.metric("Semana Actual", f"{velocidad_semana_actual:.0f} PAGO")
                    
                    # Segunda fila de métricas
                    col_sem4, col_sem5 = st.columns(2)
                    with col_sem4:
                        st.metric("🟢 Semanas Arriba del Promedio", f"{semanas_arriba}/{semanas_con_datos}")
                    with col_sem5:
                        st.metric("🔴 Semanas Bajo el Promedio", f"{semanas_bajo}/{semanas_con_datos}")
            else:
                st.info("No hay datos de PAGO por semana para este mes")
        
        # TAB 3: Análisis de Metas
        with tab_metas:
            st.markdown("##### 📊 Análisis de Cumplimiento de Metas")
            
            if asesor_seleccionado != "Todos":
                # Análisis individual
                analisis_metas = get_cumplimiento_metas_analisis(asesor_seleccionado, mes)
                
                if analisis_metas is not None:
                    # KPIs de Metas
                    col_met1, col_met2, col_met3 = st.columns(3)
                    
                    with col_met1:
                        color_cumpl = "🟢" if analisis_metas['cumplimiento_mensual'] >= 100 else "🟡" if analisis_metas['cumplimiento_mensual'] >= 80 else "🔴"
                        st.metric(
                            "Cumplimiento Mensual",
                            f"{color_cumpl} {analisis_metas['cumplimiento_mensual']:.0f}%",
                            f"Meta: {int(analisis_metas['meta_mensual'])} | Real: {int(analisis_metas['ventas_totales_todos'])}"
                        )
                    
                    with col_met2:
                        color_semanal = "🟢" if analisis_metas['cumplimiento_semanal_actual'] >= 100 else "🟡" if analisis_metas['cumplimiento_semanal_actual'] >= 80 else "🔴"
                        brecha_semanal = int(analisis_metas['progreso_semana_actual']) - int(analisis_metas['meta_semanal'])
                        brecha_text = f"Falta: {abs(brecha_semanal)}" if brecha_semanal < 0 else f"Sobra: +{brecha_semanal}"
                        st.metric(
                            "Meta Semanal vs Progreso",
                            f"{color_semanal} {analisis_metas['cumplimiento_semanal_actual']:.0f}%",
                            f"Meta: {int(analisis_metas['meta_semanal'])} | Real: {int(analisis_metas['progreso_semana_actual'])} | {brecha_text}"
                        )
                    
                    with col_met3:
                        st.metric(
                            "Información",
                            f"{int(analisis_metas['dias_trabajados'])} días",
                            f"{int(analisis_metas['semanas_activas'])} semanas activas"
                        )
                    
                    # Gráfico: Metas vs Ventas Brutas
                    df_metas_comparativa = pd.DataFrame({
                        'Tipo': ['Meta Mensual', 'Ventas Brutas (Real)'],
                        'Cantidad': [
                            int(analisis_metas['meta_mensual']), 
                            int(analisis_metas['ventas_totales_todos'])
                        ]
                    })
                    
                    # Determinar colores: azul para meta, verde si cumple/supera, rojo si no
                    color_ventas = '#10b981' if analisis_metas['ventas_totales_todos'] >= analisis_metas['meta_mensual'] else '#ef4444'
                    colores = ['#0066cc', color_ventas]
                    
                    fig_metas_comp = go.Figure(data=[
                        go.Bar(
                            x=df_metas_comparativa['Tipo'],
                            y=df_metas_comparativa['Cantidad'],
                            marker=dict(color=colores),
                            text=df_metas_comparativa['Cantidad'],
                            textposition='auto',
                            name='Cantidad'
                        )
                    ])
                    
                    fig_metas_comp.update_layout(
                        title=f"Cumplimiento de Meta - Ventas Brutas (INSTALADAS + PENDIENTES + CANCELADAS) - {mes}",
                        yaxis_title="Cantidad",
                        height=350,
                        showlegend=False,
                        plot_bgcolor='rgba(0,0,0,0)',
                        xaxis=dict(showgrid=False),
                        yaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray')
                    )
                    
                    st.plotly_chart(fig_metas_comp, use_container_width=True)
                    
                    # Mostrar status de cumplimiento
                    brecha = int(analisis_metas['ventas_totales_todos']) - int(analisis_metas['meta_mensual'])
                    col_stat1, col_stat2 = st.columns(2)
                    with col_stat1:
                        st.metric("Meta Mensual", int(analisis_metas['meta_mensual']))
                    with col_stat2:
                        if brecha >= 0:
                            st.metric("Ventas Brutas", int(analisis_metas['ventas_totales_todos']), f"✅ +{brecha} ventas (cumplió)")
                        else:
                            st.metric("Ventas Brutas", int(analisis_metas['ventas_totales_todos']), f"❌ {brecha} ventas (falta)")
                    
                    # Gráfico: Desempeño por Semana (VENTAS BRUTAS)
                    if analisis_metas['ventas_semanales']:
                        st.markdown("##### 📋 Desempeño por Semana (Ventas Brutas)")
                        
                        df_semanas = pd.DataFrame([
                            {'Semana': f"Semana {sem}", 'Ventas Brutas': ventas, 'Meta Semanal': int(analisis_metas['meta_semanal']), 
                             'Cumplimiento %': (ventas / analisis_metas['meta_semanal'] * 100) if analisis_metas['meta_semanal'] > 0 else 0}
                            for sem, ventas in sorted(analisis_metas['ventas_semanales'].items())
                        ])
                        
                        # Gráfico de barras: Ventas Brutas por semana con línea de meta
                        fig_semanas = go.Figure()
                        
                        fig_semanas.add_trace(go.Bar(
                            x=df_semanas['Semana'],
                            y=df_semanas['Ventas Brutas'],
                            marker=dict(color=['#10b981' if v >= analisis_metas['meta_semanal'] else '#ff6b6b' for v in df_semanas['Ventas Brutas']]),
                            text=df_semanas['Ventas Brutas'],
                            textposition='auto',
                            name='Ventas Brutas'
                        ))
                        
                        # Línea de meta semanal
                        fig_semanas.add_hline(
                            y=analisis_metas['meta_semanal'],
                            line_dash="dash",
                            line_color="blue",
                            annotation_text=f"Meta Semanal: {int(analisis_metas['meta_semanal'])}",
                            annotation_position="right"
                        )
                        
                        fig_semanas.update_layout(
                            title=f"Ventas Brutas por Semana - {mes}",
                            xaxis_title="Semana",
                            yaxis_title="Cantidad de Ventas Brutas",
                            height=350,
                            showlegend=False,
                            plot_bgcolor='rgba(0,0,0,0)',
                            xaxis=dict(showgrid=False),
                            yaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray')
                        )
                        
                        st.plotly_chart(fig_semanas, use_container_width=True)
                        
                        # Tabla: Desempeño por Semana detallado
                        df_semanas_display = df_semanas.copy()
                        df_semanas_display['Cumplimiento %'] = df_semanas_display['Cumplimiento %'].apply(lambda x: f"{x:.0f}%")
                        df_semanas_display['Estado'] = df_semanas_display.apply(
                            lambda x: '✅ Cumplió' if x['Cumplimiento %'].rstrip('%') >= '100' else '⚠️ En progreso' if x['Cumplimiento %'].rstrip('%') >= '80' else '❌ Bajo meta', 
                            axis=1
                        )
                        
                        st.dataframe(df_semanas_display[['Semana', 'Ventas Brutas', 'Meta Semanal', 'Cumplimiento %', 'Estado']], use_container_width=True, hide_index=True)
                    
                    # Oportunidades de Mejora Individual
                    st.markdown("##### 💡 Oportunidades de Mejora")
                    
                    brecha_mensual = int(analisis_metas['meta_mensual']) - int(analisis_metas['ventas_totales_todos'])
                    
                    if brecha_mensual > 0:
                        st.warning(f"⚠️ **Brecha Actual**: Necesita {brecha_mensual} ventas más para cumplir la meta mensual")
                        
                        dias_restantes = max(0, 30 - analisis_metas['dias_trabajados'])
                        if dias_restantes > 0:
                            ventas_diarias_necesarias = brecha_mensual / dias_restantes
                            st.info(f"📈 **Ritmo Necesario**: {ventas_diarias_necesarias:.1f} ventas/día en los {dias_restantes} días restantes")
                        else:
                            st.error(f"⏰ **Mes Finalizado**: No hay días restantes. Cumplimiento final: {analisis_metas['cumplimiento_mensual']:.0f}%")
                    else:
                        st.success(f"✅ **Meta Cumplida**: Ha superado la meta por {abs(brecha_mensual)} ventas ({analisis_metas['cumplimiento_mensual']:.0f}%)")
                else:
                    st.warning("⚠️ No hay datos de metas disponibles para este asesor")
            
            else:
                # Vista de Todos: Comparativa de todos los asesores
                st.markdown("##### 📊 Cumplimiento de Metas - Todos los Asesores")
                
                df_comp = get_comparativa_metas_todos(mes)
                
                if not df_comp.empty:
                    # KPIs de Equipo
                    oportunidades = get_oportunidades_mejora(mes)
                    
                    col_eq1, col_eq2, col_eq3, col_eq4 = st.columns(4)
                    
                    with col_eq1:
                        st.metric("Promedio Equipo", f"{oportunidades['promedio_equipo']:.0f}%", "Cumplimiento")
                    
                    with col_eq2:
                        st.metric("Asesores en Riesgo", oportunidades['asesores_en_riesgo'], f"de {oportunidades['total_asesores']}")
                    
                    with col_eq3:
                        st.metric("Asesores Excelentes", len(oportunidades['asesores_excelentes']), "100%+ cumplimiento")
                    
                    with col_eq4:
                        st.metric("Brecha Total", int(oportunidades['brecha_total']), "ventas faltantes")
                    
                    # Gráfico: Cumplimiento de todos los asesores
                    df_comp_sorted = df_comp.sort_values('Cumplimiento', ascending=True)
                    
                    fig_todos_metas = go.Figure(data=[
                        go.Bar(
                            y=df_comp_sorted['Asesor'],
                            x=df_comp_sorted['Cumplimiento'],
                            orientation='h',
                            marker=dict(color=df_comp_sorted['Color']),
                            text=df_comp_sorted['Cumplimiento'].apply(lambda x: f'{x:.0f}%'),
                            textposition='outside',
                            name='Cumplimiento'
                        )
                    ])
                    
                    fig_todos_metas.update_layout(
                        title=f"Cumplimiento de Metas Mensuales - {mes}",
                        xaxis_title="Cumplimiento (%)",
                        yaxis_title="Asesor",
                        height=600,
                        showlegend=False,
                        plot_bgcolor='rgba(0,0,0,0)',
                        xaxis=dict(showgrid=True, gridwidth=1, gridcolor='lightgray', range=[0, 130]),
                        margin=dict(l=200, r=100, t=80, b=50)
                    )
                    
                    st.plotly_chart(fig_todos_metas, use_container_width=True)
                    
                    # Tabla: Detalle de Cumplimiento
                    st.markdown("##### 📋 Detalle Completo")
                    
                    df_tabla = df_comp[['Asesor', 'Meta', 'Ventas', 'Cumplimiento', 'Estado', 'Brecha']].copy()
                    df_tabla['Cumplimiento'] = df_tabla['Cumplimiento'].apply(lambda x: f'{x:.0f}%')
                    df_tabla.columns = ['Asesor', 'Meta', 'Ventas Reales', 'Cumplimiento %', 'Estado', 'Brecha']
                    
                    st.dataframe(df_tabla, use_container_width=True, hide_index=True)
                    
                    # Oportunidades de Mejora del Equipo
                    st.markdown("##### 💡 Oportunidades de Mejora del Equipo")
                    
                    if oportunidades['asesores_en_riesgo'] > 0:
                        st.warning(f"⚠️ **Asesores en Riesgo**: {oportunidades['asesores_en_riesgo']} asesores están por debajo del 50% de cumplimiento")
                        
                        st.markdown("**Asesores Críticos:**")
                        for idx, row in oportunidades['asesores_bajo_cumplimiento'].head(5).iterrows():
                            brecha_ind = row['Brecha']
                            st.write(f"- **{row['Asesor']}**: {row['Cumplimiento']:.0f}% ({int(row['Ventas'])}/{int(row['Meta'])} | Falta: {int(brecha_ind)})")
                    
                    if len(oportunidades['asesores_excelentes']) > 0:
                        st.success(f"✅ **Asesores Excelentes**: {len(oportunidades['asesores_excelentes'])} asesores superaron la meta")
                        
                        st.markdown("**Top Performers:**")
                        for idx, row in oportunidades['asesores_excelentes'].head(3).iterrows():
                            exceso = row['Brecha']
                            st.write(f"- **{row['Asesor']}**: {row['Cumplimiento']:.0f}% ({int(row['Ventas'])}/{int(row['Meta'])} | Exceso: {int(exceso)})")
                    
                    if oportunidades['promedio_equipo'] < 80:
                        st.info(f"📈 **Recomendación**: El promedio del equipo es {oportunidades['promedio_equipo']:.0f}%. Enfoque en apoyar a los asesores rezagados")
                else:
                    st.warning("⚠️ No hay datos de metas disponibles")
        
        # FILA 3: RECOMENDACIONES PERSONALIZADAS
        st.markdown("#### 💡 Recomendaciones Personalizadas")
        
        recomendaciones = get_recomendaciones_asesor(asesor_seleccionado, kpis, mes)
        
        for rec in recomendaciones:
            if rec['tipo'] == 'crítica':
                st.error(f"**{rec['título']}**\n{rec['descripción']}\n\n✅ {rec['acción']}")
            elif rec['tipo'] == 'alta':
                st.warning(f"**{rec['título']}**\n{rec['descripción']}\n\n✅ {rec['acción']}")
            elif rec['tipo'] == 'media':
                st.info(f"**{rec['título']}**\n{rec['descripción']}\n\n✅ {rec['acción']}")
            else:
                st.success(f"**{rec['título']}**\n{rec['descripción']}\n\n✅ {rec['acción']}")
        
        # FILA 4: HISTORIAL DETALLADO
        st.markdown("#### 📋 Historial de Transacciones")
        
        df_historial = get_drive_history_by_asesor(asesor_seleccionado, mes)
        
        if not df_historial.empty:
            # Seleccionar columnas importantes para mostrar
            cols_mostrar = ['VENTA_NUM', 'FECHA', 'ESTADO', 'PAGO']
            if 'OBSERVACION' in df_historial.columns:
                cols_mostrar.append('OBSERVACION')
            
            # Asegurarse de que solo las columnas existentes se muestren
            cols_mostrar = [col for col in cols_mostrar if col in df_historial.columns]
            
            # Formatear la tabla
            df_mostrar = df_historial[cols_mostrar].copy()
            
            # Aplicar colores según estado
            def color_estado(estado):
                if estado == 'INSTALADO':
                    return '🟢 INSTALADO'
                elif estado == 'CANCELADO':
                    return '🔴 CANCELADO'
                elif estado == 'PENDIENTE':
                    return '🟡 PENDIENTE'
                else:
                    return estado
            
            df_mostrar['ESTADO'] = df_mostrar['ESTADO'].apply(color_estado)
            
            st.dataframe(
                df_mostrar,
                use_container_width=True,
                hide_index=True,
                height=300
            )
            
            st.caption(f"Total de registros: {len(df_historial)}")
        
        # FILA 5: TENDENCIAS SEMANALES
        st.markdown("#### 📆 Tendencias por Semana")
        
        tendencias = get_drive_tendencias(asesor_seleccionado, mes)
        
        if not tendencias.empty:
            col_chart, col_table = st.columns([2, 1])
            
            with col_chart:
                # Usar Plotly en lugar de st.bar_chart para evitar problemas con Altair
                tendencias_reset = tendencias.reset_index()
                fig = px.bar(tendencias_reset, x='SEMANA', y=tendencias_reset.columns.difference(['SEMANA']),
                            title='Tendencias por Semana', barmode='group')
                st.plotly_chart(fig, use_container_width=True)
            
            with col_table:
                st.dataframe(tendencias, use_container_width=True)
        else:
            st.info("No hay datos de tendencias para este asesor")
    
    else:
        st.warning(f"No hay datos del DRIVE para {asesor_seleccionado} en {mes}")

else:
    st.info(f"No hay asesores con registros en el DRIVE para {mes}")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ============= NUEVO BLOQUE: ANÁLISIS DE LEADS CON COBERTURA POR HORA Y FECHA =============
st.markdown("### 📅 Análisis de Leads con Cobertura por Hora y Fecha")
st.markdown("*Tabla que muestra la cantidad de Leads con cobertura (MANTRA) agrupados por hora del día y fecha*")

# Filtro de asesor para esta sección
col_filtro_asesor_1, col_filtro_asesor_2 = st.columns([3, 1])
with col_filtro_asesor_1:
    df_mantra_temp = load_mantra_data()
    if df_mantra_temp is not None and not df_mantra_temp.empty:
        df_mantra_temp = df_mantra_temp[df_mantra_temp['Mes'] == mes].copy()
        if not df_mantra_temp.empty:
            df_mantra_temp['Agente'] = df_mantra_temp['Agente'].astype(str).str.strip()
            opciones_asesores_cobertura = sorted(df_mantra_temp['Agente'].unique().tolist())
            asesor_filtro_cobertura = st.multiselect("👤 Filtrar por Asesor", opciones_asesores_cobertura, default=opciones_asesores_cobertura, key="asesor_cobertura")
            # Si no hay selección, considerar "Todos"
            if not asesor_filtro_cobertura:
                asesor_filtro_cobertura = "Todos"
        else:
            asesor_filtro_cobertura = "Todos"
    else:
        asesor_filtro_cobertura = "Todos"

# CREAR TABS
tab_cobertura, tab_contrato_ok = st.tabs(["📊 Leads con Cobertura", "✅ Contrato OK"])

# ============= TAB 1: LEADS CON COBERTURA =============
with tab_cobertura:
    # Obtener datos
    tabla_cobertura = get_leads_cobertura_por_hora_fecha(mes, asesor_filtro_cobertura)
    
    if not tabla_cobertura.empty:
        # Mostrar información resumida
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_leads_cobertura = tabla_cobertura.loc['TOTAL', 'TOTAL'] if 'TOTAL' in tabla_cobertura.index else 0
            st.metric("📊 Total Leads con Cobertura", int(total_leads_cobertura))
        
        with col2:
            # Encontrar la hora con más leads
            if 'TOTAL' in tabla_cobertura.index:
                tabla_sin_total = tabla_cobertura.drop('TOTAL')
                if not tabla_sin_total.empty:
                    hora_max = tabla_sin_total['TOTAL'].idxmax() if len(tabla_sin_total) > 0 else "N/A"
                    max_leads_hora = tabla_sin_total['TOTAL'].max() if len(tabla_sin_total) > 0 else 0
                    st.metric(f"⏰ Hora Pico", f"{int(hora_max)}:00 hrs", delta=f"{int(max_leads_hora)} leads")
            else:
                st.metric("⏰ Hora Pico", "N/A")
        
        with col3:
            # Promedio de leads por hora
            if 'TOTAL' in tabla_cobertura.index:
                tabla_sin_total = tabla_cobertura.drop('TOTAL')
                if len(tabla_sin_total) > 0:
                    promedio = tabla_sin_total['TOTAL'].mean()
                    st.metric("📈 Promedio por Hora", f"{promedio:.0f} leads")
            else:
                st.metric("📈 Promedio por Hora", "N/A")
    
    # Tabla detallada con estilos destacando horas pico
    st.markdown("#### 📋 Detalle Completo por Hora y Fecha")
    
    # Convertir a columnas de formato legible
    tabla_display = tabla_cobertura.copy()
    tabla_display.index.name = 'Hora'
    tabla_display = tabla_display.reset_index()
    
    # Renombrar columnas de fecha para que sean más legibles
    tabla_display.columns = ['Hora'] + [str(col).split()[0] if col != 'TOTAL' else 'TOTAL' for col in tabla_display.columns[1:]]
    
    # Crear tabla HTML con colores para horas pico - COMPACTA
    html_table_horas = '<table style="width:100%; border-collapse: collapse; margin: 5px auto; font-size: 11px;">'
    
    # Encabezado
    html_table_horas += '<tr style="background-color: #f0f2f6; border-bottom: 2px solid #ddd;">'
    for col in tabla_display.columns:
        html_table_horas += f'<th style="padding: 4px; text-align: center; border-right: 1px solid #ddd; font-size: 10px;">{col}</th>'
    html_table_horas += '</tr>'
    
    # Filas de datos
    horas_pico = [10, 11, 12, 14, 15, 16]
    for idx, row in tabla_display.iterrows():
        # Saltar la fila de TOTAL
        if str(row['Hora']).upper() == 'TOTAL':
            continue
            
        try:
            hora = int(row['Hora'])
        except (ValueError, TypeError):
            continue
            
        es_pico = hora in horas_pico
        bg_color = '#fef08a' if es_pico else '#ffffff'  # Amarillo para pico, blanco para normal
        border_left = '5px solid #ef4444' if es_pico else '1px solid #eee'  # Rojo para pico
        
        html_table_horas += f'<tr style="background-color: {bg_color}; border-bottom: 1px solid #eee; border-left: {border_left};">'
        
        for col_idx, col in enumerate(tabla_display.columns):
            valor = row[col]
            is_total = col == 'TOTAL'
            font_weight = 'bold' if is_total or es_pico else 'normal'
            color = '#ef4444' if es_pico and is_total else '#1e293b'
            
            if es_pico and col_idx == 0:
                html_table_horas += f'<td style="padding: 4px; text-align: center; border-right: 1px solid #eee; font-weight: {font_weight}; color: {color}; font-size: 11px;"><span style="background-color: #ef4444; color: white; padding: 1px 3px; border-radius: 2px; font-size: 9px;">🔥 {valor}</span></td>'
            else:
                html_table_horas += f'<td style="padding: 4px; text-align: center; border-right: 1px solid #eee; font-weight: {font_weight}; color: {color}; font-size: 11px;">{valor}</td>'
        
        html_table_horas += '</tr>'
    
    # Agregar fila de TOTAL al final
    total_row = tabla_display[tabla_display['Hora'] == 'TOTAL']
    if not total_row.empty:
        html_table_horas += '<tr style="background-color: #f0f2f6; border-top: 2px solid #ddd; font-weight: bold;">'
        for col in tabla_display.columns:
            valor = total_row[col].values[0] if col in total_row.columns else 0
            html_table_horas += f'<td style="padding: 4px; text-align: center; border-right: 1px solid #ddd; font-weight: bold; font-size: 11px;">{valor}</td>'
        html_table_horas += '</tr>'
    
    html_table_horas += '</table>'
    
    st.markdown(html_table_horas, unsafe_allow_html=True)
    
    # Análisis adicional destacando horarios pico
    st.markdown("#### 🔥 Análisis de Horarios Pico")
    
    tabla_horas = tabla_cobertura.drop('TOTAL', errors='ignore')
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Horarios Pico Identificados:**")
        # Mañana pico (10-12)
        horas_mañana_pico = [10, 11, 12]
        horas_validas_mañana = [h for h in horas_mañana_pico if h in tabla_horas.index]
        if horas_validas_mañana:
            subtotal_mañana = tabla_horas.loc[horas_validas_mañana, 'TOTAL'].sum() if 'TOTAL' in tabla_horas.columns else tabla_horas.loc[horas_validas_mañana].sum(axis=1).sum()
        else:
            subtotal_mañana = 0
        st.metric("☀️ Mañana Pico (10-12h)", int(subtotal_mañana), "leads")
    
    with col2:
        # Tarde pico (14-16)
        horas_tarde_pico = [14, 15, 16]
        horas_validas_tarde = [h for h in horas_tarde_pico if h in tabla_horas.index]
        if horas_validas_tarde:
            subtotal_tarde = tabla_horas.loc[horas_validas_tarde, 'TOTAL'].sum() if 'TOTAL' in tabla_horas.columns else tabla_horas.loc[horas_validas_tarde].sum(axis=1).sum()
        else:
            subtotal_tarde = 0
        st.metric("🌅 Tarde Pico (14-16h)", int(subtotal_tarde), "leads")
    
    # Franjas horarias generales
    st.markdown("**Desglose por Franja Horaria:**")
    
    franjas = {
        'Madrugada (0-6h)': [h for h in range(0, 7)],
        'Mañana (7-12h)': [h for h in range(7, 13)],
        'Tarde (13-18h)': [h for h in range(13, 19)],
        'Noche (19-23h)': [h for h in range(19, 24)]
    }
    
    col_franjas = st.columns(4)
    
    for idx, (franja_nombre, horas) in enumerate(franjas.items()):
        horas_validas = [h for h in horas if h in tabla_horas.index]
        if horas_validas:
            subtotal = tabla_horas.loc[horas_validas, 'TOTAL'].sum() if 'TOTAL' in tabla_horas.columns else tabla_horas.loc[horas_validas].sum(axis=1).sum()
        else:
            subtotal = 0
        
        with col_franjas[idx]:
            st.metric(franja_nombre, int(subtotal))

    # Sección de detalles: Quién se conectó en una hora-fecha específica
    st.markdown("#### 👥 Detalle: Quién se conectó")
    
    col_hora, col_fecha = st.columns(2)
    
    with col_hora:
        horas_disponibles = sorted([h for h in tabla_cobertura.index if isinstance(h, (int, float)) and h != 'TOTAL'])
        opciones_horas = ["Todos"] + horas_disponibles
        hora_detalle_idx = st.selectbox("Selecciona Hora", range(len(opciones_horas)), format_func=lambda i: str(opciones_horas[i]), key="hora_detalle")
        hora_detalle = opciones_horas[hora_detalle_idx]
    
    with col_fecha:
        fechas_disponibles = sorted([col for col in tabla_cobertura.columns if col != 'TOTAL'])
        fecha_detalle = st.selectbox("Selecciona Fecha", fechas_disponibles, key="fecha_detalle")
    
    if fecha_detalle:
        # Convertir hora_detalle a int si no es "Todos"
        hora_filtro = "Todos" if hora_detalle == "Todos" else int(hora_detalle)
        df_detalle = get_detalle_leads_cobertura(mes, asesor_filtro_cobertura, hora_filtro, str(fecha_detalle))
        
        if not df_detalle.empty:
            # Generar el mensaje según si se filtró por hora o "Todos"
            if hora_detalle == "Todos":
                titulo = f"**Leads con cobertura el {fecha_detalle} (Todas las horas)**"
            else:
                titulo = f"**Leads con cobertura el {fecha_detalle} a las {int(hora_detalle)}:00h**"
            
            st.markdown(titulo)
            
            # Agrupar por Agente y contar
            resumen_asesores = df_detalle.groupby('Agente').size().reset_index(name='Cantidad').sort_values('Cantidad', ascending=False)
            
            col_resumen, col_tabla = st.columns([1, 2])
            
            with col_resumen:
                st.markdown("**Resumen por Asesor:**")
                for idx, row in resumen_asesores.iterrows():
                    st.write(f"• {row['Agente']}: **{int(row['Cantidad'])}** leads")
            
            with col_tabla:
                st.markdown("**Detalle Completo:**")
                # Mostrar tabla con columnas limitadas
                columnas_mostrar = [col for col in df_detalle.columns if col in ['Agente', 'Dia', 'Hora']]
                st.dataframe(df_detalle[columnas_mostrar], use_container_width=True, hide_index=True)
        else:
            # Mensaje según si se filtró por hora o "Todos"
            if hora_detalle == "Todos":
                st.info(f"No hay leads en {fecha_detalle}")
            else:
                st.info(f"No hay leads en {fecha_detalle} a las {int(hora_detalle)}:00h")

    else:
        st.info(f"No hay datos de Leads con cobertura para {mes}")

# ============= TAB 2: CONTRATO OK =============
with tab_contrato_ok:
    # Obtener datos de Contrato OK
    tabla_contrato_ok = get_leads_contrato_ok_por_hora_fecha(mes, asesor_filtro_cobertura)
    
    if not tabla_contrato_ok.empty:
        # Mostrar información resumida
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_contrato_ok = tabla_contrato_ok.loc['TOTAL', 'TOTAL'] if 'TOTAL' in tabla_contrato_ok.index else 0
            st.metric("✅ Total Contrato OK", int(total_contrato_ok))
        
        with col2:
            # Encontrar la hora con más leads
            if 'TOTAL' in tabla_contrato_ok.index:
                tabla_sin_total = tabla_contrato_ok.drop('TOTAL')
                if not tabla_sin_total.empty:
                    hora_max = tabla_sin_total['TOTAL'].idxmax() if len(tabla_sin_total) > 0 else "N/A"
                    max_leads_hora = tabla_sin_total['TOTAL'].max() if len(tabla_sin_total) > 0 else 0
                    st.metric(f"⏰ Hora Pico", f"{int(hora_max)}:00 hrs", delta=f"{int(max_leads_hora)} leads")
            else:
                st.metric("⏰ Hora Pico", "N/A")
        
        with col3:
            # Promedio de leads por hora
            if 'TOTAL' in tabla_contrato_ok.index:
                tabla_sin_total = tabla_contrato_ok.drop('TOTAL')
                if len(tabla_sin_total) > 0:
                    promedio = tabla_sin_total['TOTAL'].mean()
                    st.metric("📈 Promedio por Hora", f"{promedio:.0f} leads")
            else:
                st.metric("📈 Promedio por Hora", "N/A")
        
        # Tabla detallada con estilos
        st.markdown("#### 📋 Detalle Completo por Hora y Fecha - Contrato OK")
        
        # Convertir a columnas de formato legible
        tabla_display = tabla_contrato_ok.copy()
        tabla_display.index.name = 'Hora'
        tabla_display = tabla_display.reset_index()
        
        # Renombrar columnas de fecha
        tabla_display.columns = ['Hora'] + [str(col).split()[0] if col != 'TOTAL' else 'TOTAL' for col in tabla_display.columns[1:]]
        
        # Crear tabla HTML
        html_table_contrato = '<table style="width:100%; border-collapse: collapse; margin: 5px auto; font-size: 11px;">'
        
        # Encabezado
        html_table_contrato += '<tr style="background-color: #d1fae5; border-bottom: 2px solid #ddd;">'
        for col in tabla_display.columns:
            html_table_contrato += f'<th style="padding: 4px; text-align: center; border-right: 1px solid #ddd; font-size: 10px;">{col}</th>'
        html_table_contrato += '</tr>'
        
        # Filas de datos
        for idx, row in tabla_display.iterrows():
            # Saltar la fila de TOTAL
            if str(row['Hora']).upper() == 'TOTAL':
                continue
                
            try:
                hora = int(row['Hora'])
            except (ValueError, TypeError):
                continue
            
            html_table_contrato += f'<tr style="background-color: #ffffff; border-bottom: 1px solid #eee;">'
            
            for col_idx, col in enumerate(tabla_display.columns):
                valor = row[col]
                is_total = col == 'TOTAL'
                font_weight = 'bold' if is_total else 'normal'
                color = '#059669' if is_total else '#1e293b'
                
                html_table_contrato += f'<td style="padding: 4px; text-align: center; border-right: 1px solid #eee; font-weight: {font_weight}; color: {color}; font-size: 11px;">{valor}</td>'
            
            html_table_contrato += '</tr>'
        
        # Agregar fila de TOTAL al final
        total_row = tabla_display[tabla_display['Hora'] == 'TOTAL']
        if not total_row.empty:
            html_table_contrato += '<tr style="background-color: #d1fae5; border-top: 2px solid #ddd; font-weight: bold;">'
            for col in tabla_display.columns:
                valor = total_row[col].values[0] if col in total_row.columns else 0
                html_table_contrato += f'<td style="padding: 4px; text-align: center; border-right: 1px solid #ddd; font-weight: bold; font-size: 11px;">{valor}</td>'
            html_table_contrato += '</tr>'
        
        html_table_contrato += '</table>'
        
        st.markdown(html_table_contrato, unsafe_allow_html=True)
        
        # Análisis de distribución
        st.markdown("#### 📊 Análisis de Distribución")
        
        tabla_horas_ok = tabla_contrato_ok.drop('TOTAL', errors='ignore')
        
        # Franjas horarias
        st.markdown("**Desglose por Franja Horaria:**")
        
        franjas = {
            'Madrugada (0-6h)': [h for h in range(0, 7)],
            'Mañana (7-12h)': [h for h in range(7, 13)],
            'Tarde (13-18h)': [h for h in range(13, 19)],
            'Noche (19-23h)': [h for h in range(19, 24)]
        }
        
        col_franjas = st.columns(4)
        
        for idx, (franja_nombre, horas) in enumerate(franjas.items()):
            horas_validas = [h for h in horas if h in tabla_horas_ok.index]
            if horas_validas:
                subtotal = tabla_horas_ok.loc[horas_validas, 'TOTAL'].sum() if 'TOTAL' in tabla_horas_ok.columns else tabla_horas_ok.loc[horas_validas].sum(axis=1).sum()
            else:
                subtotal = 0
            
            with col_franjas[idx]:
                st.metric(franja_nombre, int(subtotal))
    
    else:
        st.info(f"No hay datos de Contrato OK para {mes}")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Tabla de resumen mensual con expanders
st.markdown("### ⭐ Resumen Mensual Completo")

st.markdown('<div style="margin: 20px 0;"></div>', unsafe_allow_html=True)

# Obtener datos para cada mes
meses_disponibles = get_meses_disponibles()
datos_meses = []
totales = {'Leads': 0, 'Contr': 0, 'Cober': 0}

for mes_año, mes_nombre, año, mes_num in meses_disponibles:
    leads, conversion = get_total_leads_and_conversion(mes_nombre)
    con_cobertura = get_con_cobertura_count(mes_nombre)
    cancelados = get_cancelados_mes(mes_nombre)
    instaladas = get_instaladas_mes(mes_nombre)
    no_pago = get_no_pago_mes(mes_nombre)
    no_responde = get_no_responde_mes(mes_nombre)
    no_especifica = get_no_especifica_mes(mes_nombre)
    sin_cobertura = get_sin_cobertura_mes(mes_nombre)
    datos_meses.append({
        'Mes': mes_nombre,
        'Año': año,
        'Mes_Año': f'{mes_nombre} {año}',
        'Leads': leads,
        'Cober': con_cobertura,
        'Contr': conversion,
        'Cancel': cancelados,
        'Pago': instaladas,
        'NoPago': no_pago,
        'NoResp': no_responde,
        'NoEsp': no_especifica,
        'SinCob': sin_cobertura
    })
    totales['Leads'] += leads
    totales['Cober'] += con_cobertura
    totales['Contr'] += conversion

# Ordenar datos por año y número de mes (para que aparezcan en orden cronológico correcto)
datos_meses.sort(key=lambda x: (x['Año'], [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
].index(x['Mes'])))

# Construir tabla HTML de resumen sin expanders (solo datos de resumen)
html_resumen = '''
<div class="resumen-tabla">
<table>
<thead><tr>
<th>Mes</th>
<th>Leads</th>
<th>Cober</th>
<th>%Cob</th>
<th>Contr</th>
<th>%Conv</th>
<th>Cancel</th>
<th>Pagó</th>
<th>NoPag</th>
<th>Efect</th>
<th>NoResp</th>
<th>%NR</th>
<th>NoEsp</th>
<th>%NE</th>
<th>%SC</th>
</tr></thead><tbody>
'''

for dato in datos_meses:
    mes_nombre = dato['Mes']
    año = dato['Año']
    mes_año_display = dato['Mes_Año']
    leads = dato['Leads']
    cober = dato['Cober']
    cob_pct = int(cober/leads*100) if leads > 0 else 0
    contr = dato['Contr']
    conv_pct = get_conversion_mantra_mes(mes_nombre)
    cancel = dato['Cancel']
    pago = dato['Pago']
    nopago = dato['NoPago']
    efect_pct = int(pago/(cancel+pago+nopago)*100) if (cancel+pago+nopago) > 0 else 0
    noresp = dato['NoResp']
    noresp_pct = int(noresp/leads*100) if leads > 0 else 0
    noesp = dato['NoEsp']
    noesp_pct = int(noesp/leads*100) if leads > 0 else 0
    sincob = dato['SinCob']
    sincob_pct = int(sincob/leads*100) if leads > 0 else 0
    
    html_resumen += f'''<tr>
    <td><strong>{mes_año_display}</strong></td>
    <td>{leads}</td>
    <td>{cober}</td>
    <td>{cob_pct}%</td>
    <td>{contr}</td>
    <td style="color: #0066cc; font-weight: 700;">{conv_pct}%</td>
    <td>{cancel}</td>
    <td>{pago}</td>
    <td>{nopago}</td>
    <td style="color: #0066cc; font-weight: 700;">{efect_pct}%</td>
    <td>{noresp}</td>
    <td>{noresp_pct}%</td>
    <td>{noesp}</td>
    <td>{noesp_pct}%</td>
    <td>{sincob_pct}%</td>
    </tr>'''

html_resumen += '</tbody></table></div>'

st.markdown(html_resumen, unsafe_allow_html=True)

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Footer elegante
st.markdown("""
<div class="footer-container">
    <p><strong>Dashboard WORLD TEL</strong> - Sistema de Control de Cumplimiento Mensual</p>
    <p>📅 Periodo: Noviembre 2025 | 🕐 Actualizado: {}  | 👥 Total Empleados: 14</p>
    <p style="margin-top: 15px; opacity: 0.7;">© 2025 WORLD TEL | Todos los derechos reservados</p>
</div>
""".format(datetime.now().strftime("%d/%m/%Y %H:%M:%S")), unsafe_allow_html=True)
